import os
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from supabase import create_client

def load_env():
    env = {}
    if os.path.exists(".env"):
        with open(".env", "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    k, v = line.split("=", 1)
                    env[k.strip()] = v.strip()
    return env

env = load_env()
url = "https://dejlrdtreenbaquwuyxq.supabase.co"
key = env.get("SUPABASE_SERVICE_ROLE_KEY") or os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImRlamxyZHRyZWVuYmFxdXd1eXhxIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc4MTYxNDAwMSwiZXhwIjoyMDk3MTkwMDAxfQ.Dz8jLp9cvMtLA0XVAZXJsz1BjI0ctuCj0gudl4W5sUY"

supabase = create_client(url, key)

DELITOS = [
  'Robo a personas',
  'Robo a unidades económicas',
  'Robo a domicilios',
]
ALL_VARS = DELITOS
PERIODOS = ['2023', '2024', '2025', '2026*']

def parse_inauguracion_year(proj_name, date_str):
    # User requested override
    if proj_name in ["Calle Rocafuerte", "Escalinatas Rocafuerte"]:
        return 2025, "11 de Diciembre de 2025"
    if not date_str:
        return None, None
    d_clean = date_str.strip().lower()
    if d_clean in ["no", "reprogramado", "", "-", "n/a", ""]:
        return None, None
    if "2023" in d_clean:
        return 2023, date_str
    elif "2024" in d_clean:
        return 2024, date_str
    elif "2025" in d_clean:
        return 2025, date_str
    elif "2026" in d_clean:
        return 2026, date_str
    return None, None

def main():
    print("Fetching projects from Supabase...")
    res_p = supabase.table("proyectos").select("*").execute()
    proyectos = res_p.data or []
    print(f"Fetched {len(proyectos)} projects.")

    print("Fetching security stats (paginated)...")
    seguridad = []
    page = 0
    page_size = 1000
    has_more = True
    while has_more:
        res_s = supabase.table("seguridad_estadisticas").select("*").range(page * page_size, (page + 1) * page_size - 1).execute()
        page_data = res_s.data or []
        seguridad.extend(page_data)
        has_more = len(page_data) == page_size
        page += 1
    print(f"Fetched {len(seguridad)} security records.")

    # Filter projects with inauguration dates (applying overrides)
    inaugurated_projs = []
    for p in proyectos:
        name = p['nombre']
        raw_date = p.get('fecha_inauguracion')
        year, date_text = parse_inauguracion_year(name, raw_date)
        if year is not None:
            p['inauguracion_year'] = year
            p['inauguracion_date_text'] = date_text
            inaugurated_projs.append(p)

    # Sort projects by category and then by name
    inaugurated_projs.sort(key=lambda x: (x.get('categoria') or '', x.get('nombre') or ''))
    print(f"Found {len(inaugurated_projs)} projects with inauguration dates.")

    wb = openpyxl.Workbook()
    
    # Sheet 1: Evaluacion Relativa
    ws1 = wb.active
    ws1.title = "Evaluación de Impacto"
    
    # Sheet 2: Datos por Año Calendario
    ws2 = wb.create_sheet("Histórico Año Calendario")
    
    # Style definitions
    font_title = Font(name="Segoe UI", size=15, bold=True, color="1F497D")
    font_subtitle = Font(name="Segoe UI", size=9, italic=True, color="595959")
    font_header = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=9)
    font_bold = Font(name="Segoe UI", size=9, bold=True)
    font_muted_text = Font(name="Segoe UI", size=8, italic=True, color="7F7F7F")
    
    fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    fill_sub_header = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    fill_t0 = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # soft yellow
    fill_t_minus = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid") # soft red/pink
    fill_t_plus = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # soft green
    
    border_thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # 2026 blend projection logic
    def get_projections_for_project(proj_id):
        proj_stats = [s for s in seguridad if s['proyecto_id'] == proj_id]
        proys = {}
        for var in ALL_VARS:
            v26_obs = next((s['valor'] for s in proj_stats if s['anio'] == '2026*' and s['variable'] == var), None)
            if v26_obs is None:
                proys[var] = None
                continue
            proj_scaling = (v26_obs / 4.0) * 12.0
            
            past_vals = []
            for y in ['2023', '2024', '2025']:
                val = next((s['valor'] for s in proj_stats if s['anio'] == y and s['variable'] == var), None)
                if val is not None:
                    past_vals.append(val)
            proj_prev = (sum(past_vals) / len(past_vals)) if past_vals else None
            
            blend_factor = 0.6
            if proj_prev is None:
                proys[var] = round(proj_scaling)
            else:
                proys[var] = round(blend_factor * proj_scaling + (1.0 - blend_factor) * proj_prev)
        return proys

    def get_proj_val(proj_id, var, year, proys_2026):
        if year == '2026_proy':
            return proys_2026.get(var)
        proj_stats = [s for s in seguridad if s['proyecto_id'] == proj_id]
        val = next((s['valor'] for s in proj_stats if s['anio'] == year and s['variable'] == var), None)
        return val

    def get_sum(proj_id, var_list, year, proys_2026):
        s = 0
        has_val = False
        for var in var_list:
            v = get_proj_val(proj_id, var, year, proys_2026)
            if v is not None:
                s += v
                has_val = True
        return s if has_val else None

    # --- SHEET 1: EVALUACIÓN DE IMPACTO (CALENDARIO ALINEADO) ---
    ws1.cell(row=2, column=1, value="EVALUACIÓN DE IMPACTO DE SEGURIDAD (CONSOLIDADO COMPARATIVO DE DELITOS)").font = font_title
    ws1.cell(row=3, column=1, value="Comparación de delitos un año antes de la inauguración, año de entrega y años posteriores").font = font_subtitle
    ws1.cell(row=4, column=1, value=f"Generado el {datetime.datetime.now().strftime('%d/%m/%Y')} - Fuente: Secretaría de Hábitat / PPNN").font = font_subtitle
    
    # Table headers
    ws1.merge_cells('A6:A7')
    ws1.merge_cells('B6:B7')
    ws1.merge_cells('C6:C7')
    ws1.merge_cells('D6:D7')
    
    ws1.cell(row=6, column=1, value="Nombre del Proyecto").font = font_header
    ws1.cell(row=6, column=2, value="Categoría").font = font_header
    ws1.cell(row=6, column=3, value="Fecha Inauguración").font = font_header
    ws1.cell(row=6, column=4, value="Año Inaug.").font = font_header
    
    for c in range(1, 5):
        ws1.cell(row=6, column=c).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws1.cell(row=6, column=c).fill = fill_header
    
    ws1.merge_cells('E6:L6')
    ws1.cell(row=6, column=5, value="TOTAL DELITOS").font = font_header
    ws1.cell(row=6, column=5).alignment = Alignment(horizontal='center', vertical='center')
    
    headers_sub1 = [
        "Año Pre", "Cantidad Pre (1 Año Antes)", "Año Inaug.", "Cantidad Inauguración", 
        "Cantidad 2024 (Post-Inaug.)", "Cantidad 2025 (Post-Inaug.)", "Cantidad 2026 (Proyectado)", "Tasa Cambio Total*"
    ]
    for idx, text in enumerate(headers_sub1):
        col = idx + 5
        cell_sub = ws1.cell(row=7, column=col, value=text)
        cell_sub.font = font_header
        cell_sub.fill = fill_sub_header
        cell_sub.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for r in [6, 7]:
        for c in range(1, 13):
            cell = ws1.cell(row=r, column=c)
            cell.border = border_thin
            if not cell.fill.fill_type:
                cell.fill = fill_header

    # Populate Data
    row_idx = 8
    for p in inaugurated_projs:
        name = p['nombre']
        cat = p['categoria']
        date_txt = p['inauguracion_date_text']
        t0_year = p['inauguracion_year']
        proj_id = p['id']
        
        proys_2026 = get_projections_for_project(proj_id)
        
        pre_year = t0_year - 1
        
        def get_cal_val(var_list, year_num):
            if year_num == 2023:
                return get_sum(proj_id, var_list, '2023', proys_2026)
            elif year_num == 2024:
                return get_sum(proj_id, var_list, '2024', proys_2026)
            elif year_num == 2025:
                return get_sum(proj_id, var_list, '2025', proys_2026)
            elif year_num == 2026:
                return get_sum(proj_id, var_list, '2026_proy', proys_2026)
            return None

        # Delitos values mapping
        del_pre_val = get_cal_val(DELITOS, pre_year)
        del_inaug_val = get_cal_val(DELITOS, t0_year)
        del_post_24 = get_cal_val(DELITOS, 2024) if t0_year < 2024 else None
        del_post_25 = get_cal_val(DELITOS, 2025) if t0_year < 2025 else None
        del_post_26_proy = get_cal_val(DELITOS, 2026) if t0_year < 2026 else None

        def calc_rate(pre_val, inaug_val, post_24, post_25, post_26):
            last_post = next((v for v in [post_26, post_25, post_24] if v is not None), None)
            if pre_val is not None:
                return (last_post - pre_val) / pre_val if last_post is not None else (inaug_val - pre_val) / pre_val
            elif pre_val is None and inaug_val is not None:
                return (last_post - inaug_val) / inaug_val if last_post is not None else None
            return None

        rate_del = calc_rate(del_pre_val, del_inaug_val, del_post_24, del_post_25, del_post_26_proy)

        # General Info
        ws1.cell(row=row_idx, column=1, value=name).font = font_bold
        ws1.cell(row=row_idx, column=2, value=cat).font = font_data
        ws1.cell(row=row_idx, column=3, value=date_txt).font = font_data
        ws1.cell(row=row_idx, column=4, value=t0_year).font = font_bold
        ws1.cell(row=row_idx, column=4).alignment = Alignment(horizontal='center')
        
        # Write cells for Delitos (Col E to L)
        ws1.cell(row=row_idx, column=5, value=pre_year if pre_year >= 2023 else "N/D").font = font_bold
        ws1.cell(row=row_idx, column=5).alignment = Alignment(horizontal='center')
        ws1.cell(row=row_idx, column=5).fill = fill_t_minus
        
        # Cantidad Pre
        if pre_year < 2023:
            cell_n = ws1.cell(row=row_idx, column=6, value="Sin registro (Año 2022)")
            cell_n.font = font_muted_text
            cell_n.alignment = Alignment(horizontal='center')
        else:
            ws1.cell(row=row_idx, column=6, value=del_pre_val if del_pre_val is not None else 0).font = font_data
        ws1.cell(row=row_idx, column=6).fill = fill_t_minus
        
        # Año Inaug
        ws1.cell(row=row_idx, column=7, value=t0_year).font = font_bold
        ws1.cell(row=row_idx, column=7).alignment = Alignment(horizontal='center')
        ws1.cell(row=row_idx, column=7).fill = fill_t0
        
        # Cantidad Inaug
        ws1.cell(row=row_idx, column=8, value=del_inaug_val if del_inaug_val is not None else 0).fill = fill_t0
        
        # Cantidad 2024 Post
        cell_q = ws1.cell(row=row_idx, column=9)
        if t0_year == 2024:
            cell_q.value = "Año de Inauguración (2024)"
            cell_q.font = font_muted_text
            cell_q.alignment = Alignment(horizontal='center')
        elif t0_year > 2024:
            cell_q.value = "Pre-Inauguración (2024)"
            cell_q.font = font_muted_text
            cell_q.alignment = Alignment(horizontal='center')
        else:
            cell_q.value = del_post_24 if del_post_24 is not None else 0
            cell_q.font = font_data
        cell_q.fill = fill_t_plus
        
        # Cantidad 2025 Post
        cell_r = ws1.cell(row=row_idx, column=10)
        if t0_year == 2025:
            cell_r.value = "Año de Inauguración (2025)"
            cell_r.font = font_muted_text
            cell_r.alignment = Alignment(horizontal='center')
        elif t0_year > 2025:
            cell_r.value = "Pre-Inauguración (2025)"
            cell_r.font = font_muted_text
            cell_r.alignment = Alignment(horizontal='center')
        else:
            cell_r.value = del_post_25 if del_post_25 is not None else 0
            cell_r.font = font_data
        cell_r.fill = fill_t_plus
        
        # Cantidad 2026 Proy
        cell_s = ws1.cell(row=row_idx, column=11)
        if t0_year == 2026:
            cell_s.value = "Año de Inauguración (2026)"
            cell_s.font = font_muted_text
            cell_s.alignment = Alignment(horizontal='center')
        else:
            if del_post_26_proy is not None:
                cell_s.value = del_post_26_proy
                cell_s.font = font_data
            else:
                cell_s.value = "Sin registro (Año 2026)"
                cell_s.font = font_muted_text
                cell_s.alignment = Alignment(horizontal='center')
        cell_s.fill = fill_t_plus
        
        # Tasa de cambio
        cell_rate_del = ws1.cell(row=row_idx, column=12, value=rate_del if rate_del is not None else "N/A")
        cell_rate_del.alignment = Alignment(horizontal='center')
        if rate_del is not None:
            cell_rate_del.number_format = '+0.0%;-0.0%;0.0%'
            cell_rate_del.font = Font(name="Segoe UI", size=9, bold=True, color="C00000" if rate_del > 0 else "385723")
        else:
            cell_rate_del.font = font_data

        # Zebra coloring on first 3 cols
        if row_idx % 2 == 1:
            for c in [1, 2, 3]:
                ws1.cell(row=row_idx, column=c).fill = fill_zebra
                
        for c in range(1, 13):
            cell = ws1.cell(row=row_idx, column=c)
            cell.border = border_thin
            if c in [6, 8, 9, 10, 11] and isinstance(cell.value, int):
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0'

        row_idx += 1

    # Footnotes
    ws1.cell(row=row_idx+1, column=1, value="* Notas explicativas de la tabla:").font = font_bold
    ws1.cell(row=row_idx+2, column=1, value="  - Año Pre (Rojo): Corresponde a exactamente un año calendario antes de la fecha de inauguración (Periodo de control Pre-Proyecto).").font = font_subtitle
    ws1.cell(row=row_idx+3, column=1, value="  - Cantidad Inauguración (Amarillo): Suma registrada durante el año de entrega del proyecto (Hito cero).").font = font_subtitle
    ws1.cell(row=row_idx+4, column=1, value="  - Cantidades por Años Posteriores (Verde): Sumas correspondientes únicamente a los años posteriores a la inauguración. Si una casilla está explicada con texto es porque ese año calendario fue la inauguración o el periodo previo del proyecto.").font = font_subtitle
    ws1.cell(row=row_idx+5, column=1, value="  - Tasa de Cambio: Mide la diferencia porcentual entre el Año Pre y el último año Post disponible (2026 Proyectado). Para proyectos de 2023, al no haber datos de 2022 (t-1), se mide desde el año de inauguración (2023).").font = font_subtitle

    # --- SHEET 2: HISTÓRICO AÑO CALENDARIO ---
    ws2.cell(row=2, column=1, value="HISTÓRICO GENERAL POR AÑO CALENDARIO (DELITOS)").font = font_title
    ws2.cell(row=3, column=1, value="Consolidado de delitos reales (2023-2025, 2026 Ene-Abr) y proyectados (2026) organizados cronológicamente").font = font_subtitle
    ws2.cell(row=4, column=1, value=f"Generado el {datetime.datetime.now().strftime('%d/%m/%Y')}").font = font_subtitle
    
    # Headers
    ws2.merge_cells('A6:A7')
    ws2.merge_cells('B6:B7')
    ws2.merge_cells('C6:C7')
    ws2.merge_cells('D6:D7')
    
    ws2.cell(row=6, column=1, value="Nombre del Proyecto").font = font_header
    ws2.cell(row=6, column=2, value="Categoría").font = font_header
    ws2.cell(row=6, column=3, value="Fecha Inauguración").font = font_header
    ws2.cell(row=6, column=4, value="Año Inaug.").font = font_header
    
    for c in range(1, 5):
        ws2.cell(row=6, column=c).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws2.cell(row=6, column=c).fill = fill_header
        
    ws2.merge_cells('E6:I6')
    ws2.cell(row=6, column=5, value="TOTAL DELITOS").font = font_header
    ws2.cell(row=6, column=5).alignment = Alignment(horizontal='center', vertical='center')
    
    headers_sub2 = [
        "2023", "2024", "2025", "2026 (Real Ene-Abr)", "2026 (Proy)"
    ]
    for idx, text in enumerate(headers_sub2):
        col = idx + 5
        cell_sub = ws2.cell(row=7, column=col, value=text)
        cell_sub.font = font_header
        cell_sub.fill = fill_sub_header
        cell_sub.alignment = Alignment(horizontal='center', vertical='center')

    for r in [6, 7]:
        for c in range(1, 10):
            cell = ws2.cell(row=r, column=c)
            cell.border = border_thin
            if not cell.fill.fill_type:
                cell.fill = fill_header

    # Populate Data
    row_idx_cal = 8
    for p in inaugurated_projs:
        name = p['nombre']
        cat = p['categoria']
        date_txt = p['inauguracion_date_text']
        t0_year = p['inauguracion_year']
        proj_id = p['id']
        
        proys_2026 = get_projections_for_project(proj_id)
        
        del_23 = get_sum(proj_id, DELITOS, '2023', proys_2026)
        del_24 = get_sum(proj_id, DELITOS, '2024', proys_2026)
        del_25 = get_sum(proj_id, DELITOS, '2025', proys_2026)
        del_26_real = get_sum(proj_id, DELITOS, '2026*', proys_2026)
        del_26_proy = get_sum(proj_id, DELITOS, '2026_proy', proys_2026)

        ws2.cell(row=row_idx_cal, column=1, value=name).font = font_bold
        ws2.cell(row=row_idx_cal, column=2, value=cat).font = font_data
        ws2.cell(row=row_idx_cal, column=3, value=date_txt).font = font_data
        ws2.cell(row=row_idx_cal, column=4, value=t0_year).font = font_bold
        ws2.cell(row=row_idx_cal, column=4).alignment = Alignment(horizontal='center')
        
        def set_cal_cell(sheet, r, c, val, is_t0):
            cell = sheet.cell(row=r, column=c, value=val if val is not None else 0)
            cell.alignment = Alignment(horizontal='right')
            if is_t0:
                cell.fill = fill_t0
                cell.font = font_bold
            else:
                cell.font = font_data
            if isinstance(val, int):
                cell.number_format = '#,##0'

        # Set values for Delitos (Col E-I)
        set_cal_cell(ws2, row_idx_cal, 5, del_23, t0_year == 2023)
        set_cal_cell(ws2, row_idx_cal, 6, del_24, t0_year == 2024)
        set_cal_cell(ws2, row_idx_cal, 7, del_25, t0_year == 2025)
        set_cal_cell(ws2, row_idx_cal, 8, del_26_real, False)
        set_cal_cell(ws2, row_idx_cal, 9, del_26_proy, t0_year == 2026)

        # Zebra striping on first 3 cols
        if row_idx_cal % 2 == 1:
            for c in [1, 2, 3]:
                if not ws2.cell(row=row_idx_cal, column=c).fill.fill_type:
                    ws2.cell(row=row_idx_cal, column=c).fill = fill_zebra
                
        for c in range(1, 10):
            ws2.cell(row=row_idx_cal, column=c).border = border_thin

        row_idx_cal += 1

    ws2.cell(row=row_idx_cal+1, column=1, value="* Nota: Las celdas con fondo amarillo indican el año calendario de inauguración (t0) de cada proyecto.").font = font_subtitle

    # Auto-adjust column widths
    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.row < 5:  # skip titles
                    continue
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 11)
            
        ws.column_dimensions['A'].width = 38
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 24

    # Save to local workspace
    filename = "Reporte_Evaluacion_Proyectos_Inaugurados_v6.xlsx"
    wb.save(filename)
    print(f"Excel saved to workspace: {filename}")
    
    # Copy to artifacts directory
    artifacts_dir = "C:\\Users\\User\\.gemini\\antigravity-ide\\brain\\f325f5d9-11bb-4a9e-bde4-48d5b9976a98"
    if os.path.exists(artifacts_dir):
        wb.save(os.path.join(artifacts_dir, filename))
        print(f"Excel copied to artifacts: {os.path.join(artifacts_dir, filename)}")

if __name__ == '__main__':
    main()
