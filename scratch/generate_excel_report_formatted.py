import os
import re
import unicodedata
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Rutas
DATA_DIR = r"C:\Users\User\Desktop\URU-Estadisticas\URU_Estadistico\data"
PREDIOS_PATH = os.path.join(DATA_DIR, 'PREDIOS.xlsx')
BDD_22_PATH = os.path.join(DATA_DIR, 'BDD LUAE 2022- Proyectos estrategicos _ Desarrollo Urbanistico.xlsx')
BDD_23_PATH = os.path.join(DATA_DIR, 'BDD- Proyectos estrategicos _ Desarrollo Urbanistico.xlsx')
OUTPUT_PATH = r"C:\Users\User\Desktop\Reporte_LUAE_Rocafuerte_Ene_Abr_Proyecciones_v2.xlsx"

def normalize(text):
    if not text or pd.isna(text):
        return ''
    s = str(text).strip().lower()
    s = ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))
    s = s.replace('\ufffd', 'o')
    s = re.sub(r'[^a-z0-9]+', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()

def clean_predio(p):
    if pd.isna(p):
        return ''
    s = str(p).strip()
    if s.endswith('.0'):
        s = s[:-2]
    return s

def norm_mov(m):
    m_norm = normalize(m)
    if 'renov' in m_norm:
        return 'RENOVACION'
    if 'emis' in m_norm:
        return 'EMISION'
    return m_norm.upper()

def main():
    print("=== EXTRAYENDO DATOS ===")
    
    # 1. Obtener predios de Calle Rocafuerte (columna 39)
    df_predios = pd.read_excel(PREDIOS_PATH, header=None)
    raw_predios = df_predios.iloc[2:, 39].dropna().tolist()
    rocafuerte_predios = set([clean_predio(p) for p in raw_predios if clean_predio(p)])
    
    # 2. Cargar y filtrar BDDs
    all_matched = []
    bdd_files = [(BDD_22_PATH, "2022"), (BDD_23_PATH, "2023+")]
    
    for path, label in bdd_files:
        df = pd.read_excel(path)
        predio_col = [c for c in df.columns if normalize(c) == 'predio'][0]
        licencia_col = [c for c in df.columns if 'licencia' in normalize(c)][0]
        mov_col = [c for c in df.columns if 'tipo de movimiento' in normalize(c) or 'movimiento' in normalize(c)][0]
        ciiu_col = [c for c in df.columns if 'ciiu' in normalize(c) and 'descrip' in normalize(c)][0]
        fecha_col = [c for c in df.columns if 'impresion' in normalize(c) or 'impresio' in normalize(c)][0]
        
        for idx, row in df.iterrows():
            p_clean = clean_predio(row[predio_col])
            if p_clean in rocafuerte_predios:
                lic = str(row[licencia_col]).strip()
                mov = str(row[mov_col]).strip()
                ciiu = str(row[ciiu_col]).strip() if pd.notna(row[ciiu_col]) else ''
                fecha_val = row[fecha_col]
                
                mov_n = norm_mov(mov)
                if mov_n not in ['EMISION', 'RENOVACION']:
                    continue
                    
                date_obj = None
                if pd.notna(fecha_val):
                    try:
                        date_obj = pd.to_datetime(fecha_val)
                    except Exception:
                        pass
                        
                all_matched.append({
                    'PREDIO': p_clean,
                    'LICENCIA': lic,
                    'MOVIMIENTO': mov_n,
                    'DESCRIPCION_CIIU': ciiu,
                    'FECHA_IMPRESION': date_obj,
                    'ANIO': date_obj.year if date_obj else None,
                    'MES': date_obj.month if date_obj else None
                })
                
    df_matched = pd.DataFrame(all_matched)
    
    # Deduplicación
    df_matched['lic_norm'] = df_matched['LICENCIA'].str.strip().str.lower()
    df_matched['predio_norm'] = df_matched['PREDIO'].str.strip().str.lower()
    df_matched['mov_norm'] = df_matched['MOVIMIENTO'].str.strip().str.upper()
    df_matched['ciiu_norm'] = df_matched['DESCRIPCION_CIIU'].str.strip().str.lower()
    df_dedup = df_matched.drop_duplicates(subset=['lic_norm', 'predio_norm', 'mov_norm', 'ciiu_norm']).copy()
    
    # Filtrar registros con año válido
    df_valid = df_dedup.dropna(subset=['ANIO', 'MES']).copy()
    df_valid['ANIO'] = df_valid['ANIO'].astype(int)
    df_valid['MES'] = df_valid['MES'].astype(int)
    
    # 3. Generar hojas de cálculo
    
    # Tab 1: Histórico Ene-Abr (2022-2026)
    ja_rows = []
    yearly_totals = {2022: 112, 2023: 54, 2024: 65, 2025: 52, 2026: '—'}
    for yr in range(2022, 2027):
        df_yr_ja = df_valid[(df_valid['ANIO'] == yr) & (df_valid['MES'] <= 4)]
        emisiones = (df_yr_ja['MOVIMIENTO'] == 'EMISION').sum()
        renovaciones = (df_yr_ja['MOVIMIENTO'] == 'RENOVACION').sum()
        ja_rows.append({
            'Año': yr,
            'Periodo': 'Enero - Abril',
            'Emisiones (Emitidas)': emisiones,
            'Renovaciones (Renovadas)': renovaciones,
            'Total LUAEs (Ene-Abr)': emisiones + renovaciones,
            'Total Anual Real (Ene-Dic)': yearly_totals[yr]
        })
    df_ja = pd.DataFrame(ja_rows)
    
    # Tab 2: Proyecciones Resto 2026 (Mayo-Diciembre)
    # Ene-Abr 2026 Real es: Emisiones = 11, Renovaciones = 11, Total = 22
    # May-Dec 2026 Proyectado:
    # Escenario 1 (Proporción Estacional): 55 total -> 34 emisiones, 21 renovaciones
    # Escenario 2 (Tendencia Reciente): 39 total -> 24 emisiones, 15 renovaciones
    # Escenario 3 (Promedio Histórico): 51 total -> 31 emisiones, 20 renovaciones
    proj_rows = [
        {
            'Escenario de Proyección': 'Escenario 1: Proporción Estacional Histórica (Jan-Apr representa 28.6% del año)',
            'Real (Ene-Abr)': 22,
            'Proyectado Emisiones (May-Dic)': 34,
            'Proyectado Renovaciones (May-Dic)': 21,
            'Total Proyectado (May-Dic)': 55,
            'Total Anual 2026 Proyectado': 77,
            'Fórmula de Proyección': 'Proyectado total = (Ene-Abr 2026 / 28.62%). El resto (71.4%) se reparte 61.4% emisiones y 38.6% renovaciones.'
        },
        {
            'Escenario de Proyección': 'Escenario 2: Tendencia Reciente (Promedio de Mayo-Diciembre 2023-2025)',
            'Real (Ene-Abr)': 22,
            'Proyectado Emisiones (May-Dic)': 24,
            'Proyectado Renovaciones (May-Dic)': 15,
            'Total Proyectado (May-Dic)': 39,
            'Total Anual 2026 Proyectado': 61,
            'Fórmula de Proyección': 'Proyectado resto = Promedio(2023, 2024, 2025) de Mayo-Diciembre. Se excluye 2022 por ser atípico.'
        },
        {
            'Escenario de Proyección': 'Escenario 3: Promedio Histórico Total (Promedio de Mayo-Diciembre 2022-2025)',
            'Real (Ene-Abr)': 22,
            'Proyectado Emisiones (May-Dic)': 31,
            'Proyectado Renovaciones (May-Dic)': 20,
            'Total Proyectado (May-Dic)': 51,
            'Total Anual 2026 Proyectado': 73,
            'Fórmula de Proyección': 'Proyectado resto = Promedio(2022, 2023, 2024, 2025) de Mayo-Diciembre.'
        }
    ]
    df_proj = pd.DataFrame(proj_rows)
    
    # Tab 3: Historial Mensual
    monthly_rows = []
    month_names = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo', 6: 'Junio',
        7: 'Julio', 8: 'Agosto', 9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
    }
    for yr in range(2022, 2027):
        for m in range(1, 13):
            if yr == 2026 and m > 4:
                continue
            df_sub = df_valid[(df_valid['ANIO'] == yr) & (df_valid['MES'] == m)]
            emisiones = (df_sub['MOVIMIENTO'] == 'EMISION').sum()
            renovaciones = (df_sub['MOVIMIENTO'] == 'RENOVACION').sum()
            monthly_rows.append({
                'Año': yr,
                'Mes': month_names[m],
                'Emisiones': emisiones,
                'Renovaciones': renovaciones,
                'Total': emisiones + renovaciones
            })
    df_monthly = pd.DataFrame(monthly_rows)
    
    # 4. Crear libro openpyxl con estilos
    wb = openpyxl.Workbook()
    
    # Fuentes y rellenos
    font_family = "Segoe UI"
    font_header = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    font_data = Font(name=font_family, size=10)
    font_bold = Font(name=font_family, size=10, bold=True)
    font_title = Font(name=font_family, size=14, bold=True, color="24367F")
    
    fill_header = PatternFill(start_color="24367F", end_color="24367F", fill_type="solid") # Dark Slate Blue de la web
    fill_zebra = PatternFill(start_color="F4F6FB", end_color="F4F6FB", fill_type="solid") # Light Blue-Gray
    fill_highlight = PatternFill(start_color="E6ECF8", end_color="E6ECF8", fill_type="solid") # Highlight azul claro para totales o proyectados
    
    border_thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    border_total = Border(
        top=Side(style='thin', color='24367F'),
        bottom=Side(style='double', color='24367F')
    )
    
    # ----------------------------------------------------
    # HOJA 1: Resumen Ene-Abr
    # ----------------------------------------------------
    ws1 = wb.active
    ws1.title = "Histórico Ene-Abr"
    ws1.views.sheetView[0].showGridLines = True
    
    ws1.append([])
    ws1.append(["Historial de Emisiones y Renovaciones (Periodo Enero - Abril)"])
    ws1.cell(2, 1).font = font_title
    ws1.append(["Proyecto: Calle Rocafuerte"])
    ws1.cell(3, 1).font = Font(name=font_family, size=10, italic=True)
    ws1.append([])
    
    # Cabeceras
    headers_1 = df_ja.columns.tolist()
    ws1.append(headers_1)
    for col_idx in range(1, len(headers_1) + 1):
        cell = ws1.cell(5, col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Datos
    for idx, row in df_ja.iterrows():
        ws1.append(row.tolist())
        row_idx = ws1.max_row
        is_even = (row_idx % 2 == 0)
        
        for col_idx in range(1, len(headers_1) + 1):
            cell = ws1.cell(row_idx, col_idx)
            cell.font = font_data
            cell.border = border_thin
            
            # Alineación
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 2:
                cell.alignment = Alignment(horizontal="left")
            elif col_idx == 6 and cell.value == '—':
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="right")
                if cell.value != '—':
                    cell.number_format = "#,##0"
                
            # Zebra striping
            if is_even:
                cell.fill = fill_zebra
                
    # Fila de Totales
    total_row = ["Total Acumulado", "", df_ja.iloc[:, 2].sum(), df_ja.iloc[:, 3].sum(), df_ja.iloc[:, 4].sum(), 283]
    ws1.append(total_row)
    last_row_idx = ws1.max_row
    for col_idx in range(1, len(total_row) + 1):
        cell = ws1.cell(last_row_idx, col_idx)
        cell.font = font_bold
        cell.border = border_total
        if col_idx >= 3:
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = "#,##0"
            
    # Auto-fit
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    # ----------------------------------------------------
    # HOJA 2: Proyecciones 2026
    # ----------------------------------------------------
    ws2 = wb.create_sheet(title="Proyecciones Resto 2026")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2.append([])
    ws2.append(["Proyección de LUAEs para el resto del año 2026 (Mayo - Diciembre)"])
    ws2.cell(2, 1).font = font_title
    ws2.append(["Proyecto: Calle Rocafuerte | Cifras proyectadas basadas en comportamiento histórico"])
    ws2.cell(3, 1).font = Font(name=font_family, size=10, italic=True)
    ws2.append([])
    
    headers_2 = df_proj.columns.tolist()
    ws2.append(headers_2)
    for col_idx in range(1, len(headers_2) + 1):
        cell = ws2.cell(5, col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Altura de cabecera de proyecciones
    ws2.row_dimensions[5].height = 28
    
    # Datos
    for idx, row in df_proj.iterrows():
        ws2.append(row.tolist())
        row_idx = ws2.max_row
        
        for col_idx in range(1, len(headers_2) + 1):
            cell = ws2.cell(row_idx, col_idx)
            cell.border = border_thin
            
            # Estilos específicos por columna
            if col_idx == 1:
                cell.font = font_bold
                cell.alignment = Alignment(horizontal="left", wrap_text=True)
            elif col_idx == 7:
                cell.font = Font(name=font_family, size=9, italic=True)
                cell.alignment = Alignment(horizontal="left", wrap_text=True)
            else:
                cell.font = font_data
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0"
                if col_idx == 6: # Destacar la columna total anual
                    cell.font = font_bold
                    cell.fill = fill_highlight
                    
    # Auto-fit para Hoja 2 con anchos fijos razonables para columnas largas
    ws2.column_dimensions['A'].width = 38
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 25
    ws2.column_dimensions['D'].width = 25
    ws2.column_dimensions['E'].width = 20
    ws2.column_dimensions['F'].width = 22
    ws2.column_dimensions['G'].width = 45
    
    # ----------------------------------------------------
    # HOJA 3: Historial Mensual
    # ----------------------------------------------------
    ws3 = wb.create_sheet(title="Historial Mensual Completo")
    ws3.views.sheetView[0].showGridLines = True
    
    ws3.append([])
    ws3.append(["Base de Datos Mensual Detallada (2022 - 2026)"])
    ws3.cell(2, 1).font = font_title
    ws3.append(["Proyecto: Calle Rocafuerte"])
    ws3.cell(3, 1).font = Font(name=font_family, size=10, italic=True)
    ws3.append([])
    
    headers_3 = df_monthly.columns.tolist()
    ws3.append(headers_3)
    for col_idx in range(1, len(headers_3) + 1):
        cell = ws3.cell(5, col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    for idx, row in df_monthly.iterrows():
        ws3.append(row.tolist())
        row_idx = ws3.max_row
        is_even = (row_idx % 2 == 0)
        
        for col_idx in range(1, len(headers_3) + 1):
            cell = ws3.cell(row_idx, col_idx)
            cell.font = font_data
            cell.border = border_thin
            
            if col_idx in [1, 2]:
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0"
                
            if is_even:
                cell.fill = fill_zebra
                
    # Auto-fit
    for col in ws3.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws3.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    # Guardar
    wb.save(OUTPUT_PATH)
    print(f"Excel generado exitosamente en: {OUTPUT_PATH}")

if __name__ == '__main__':
    main()
