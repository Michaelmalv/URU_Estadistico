import io
import os
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path
import pandas as pd
import openpyxl

# Intentar importar Supabase, si no está instalado, dar instrucciones claras
try:
    from supabase import create_client, Client
except ImportError:
    print("Error: No se ha encontrado el SDK de Supabase.")
    print("Por favor, ejecuta: pip install supabase")
    exit(1)

# Configuración de Rutas
BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / 'data'

# Variables del Excel de Seguridad
INCIDENTES = [
    'Daño a propiedad pública y privada',
    'Escándalos',
    'Eventos clandestinos',
    'Libadores',
    'Venta y consumo de sustancias',
]
DELITOS = [
    'Robo a carros', 'Robo a motos', 'Robo a personas',
    'Robo a unidades económicas', 'Robo de autopartes', 'Robo a domicilios',
]
ALL_VARS = INCIDENTES + DELITOS
COL_INICIO = {'2023': 5, '2024': 16, '2025': 27, '2026*': 38}

MONTHS_ES = {
    'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4, 'mayo': 5, 'junio': 6,
    'julio': 7, 'agosto': 8, 'septiembre': 9, 'setiembre': 9, 'octubre': 10,
    'noviembre': 11, 'diciembre': 12
}

# Funciones de normalización de datos
def safe_float(v):
    if v is None or str(v).strip() in ('', '-'):
        return None
    try:
        return float(v)
    except:
        return None

def normalize_text(value):
    if value is None:
        return ''
    text = str(value).strip().lower()
    text = unicodedata.normalize('NFKD', text).encode('ascii', 'ignore').decode('ascii')
    text = re.sub(r'[^a-z0-9]+', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    return text

def parse_flexible_date(value):
    if value is None:
        return None
    if isinstance(value, pd.Timestamp):
        value = value.to_pydatetime()
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and not pd.isna(value):
        year = int(value)
        if 1900 <= year <= 2100:
            return date(year, 1, 1)
        return None

    text = str(value).strip()
    if not text or normalize_text(text) in ('no', 'nan', 'none', '-', '—'):
        return None

    dt = pd.to_datetime(text, dayfirst=True, errors='coerce')
    if pd.notna(dt):
        return dt.date()

    norm = normalize_text(text)
    m = re.search(r'(?P<day>\d{1,2})\s*(?:de\s*)?(?P<month>[a-zñ]+)\s*(?:de\s*)?(?P<year>\d{4})', norm)
    if m:
        month = MONTHS_ES.get(m.group('month'))
        if month:
            year = int(m.group('year'))
            day = int(m.group('day'))
            try:
                return date(year, month, day)
            except ValueError:
                pass

    m = re.search(r'(?P<month>[a-zñ]+)\s+(?P<year>\d{4})', norm)
    if m:
        month = MONTHS_ES.get(m.group('month'))
        if month:
            year = int(m.group('year'))
            try:
                return date(year, month, 1)
            except ValueError:
                pass

    m = re.search(r'\b(19|20)\d{2}\b', norm)
    if m:
        year = int(m.group(0))
        return date(year, 1, 1)

    return None

def find_header_row(rows, candidates, limit=30):
    for idx, row in enumerate(rows[:limit]):
        normalized = ' | '.join(normalize_text(cell) for cell in row if cell is not None)
        if any(normalize_text(candidate) in normalized for candidate in candidates):
            return idx
    return None

def match_column_name(headers, candidates):
    normalized_headers = [normalize_text(h) for h in headers]
    normalized_candidates = [normalize_text(c) for c in candidates]
    for candidate in normalized_candidates:
        for idx, header in enumerate(normalized_headers):
            if header == candidate or candidate in header or header in candidate:
                return idx
    return None

# Importar matriz desde el archivo local de python
try:
    import sys
    sys.path.append(str(BASE_DIR))
    from senderos_matriz import MATRIZ, resolver_clave
except Exception as e:
    print(f"Advertencia: No se pudo cargar senderos_matriz.py: {e}. Las fichas se subirán sin vincular proyecto_id.")
    MATRIZ = []
    resolver_clave = lambda x: None

# Conexión a Supabase
def get_supabase_client() -> Client:
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or os.environ.get("SUPABASE_ANON_KEY")

    if not url:
        url = input("Ingresa la URL de tu Supabase (ej. https://xxx.supabase.co): ").strip()
    if not key:
        key = input("Ingresa tu Service Role Key / Anon Key de Supabase: ").strip()

    return create_client(url, key)

def main():
    print("=== INICIANDO INGESTA DE DATOS DE EXCEL A SUPABASE ===")
    supabase = get_supabase_client()

    # --- 1. PROCESAR Y CARGAR PROYECTOS Y SEGURIDAD ---
    print("\n[1/4] Procesando Excel de Seguridad...")
    excel_seguridad = DATA_DIR / 'Evaluación de proyectos estratégicos_ SEGURIDAD.xlsx'
    if not excel_seguridad.is_file():
        # Fallback a buscar cualquier xlsx en data
        files = sorted(DATA_DIR.glob('*.xlsx'))
        if files:
            excel_seguridad = files[0]
            print(f"-> Usando archivo de respaldo: {excel_seguridad.name}")
        else:
            print("Error: No se encontró el archivo de Excel de Seguridad en data/")
            return

    wb = openpyxl.load_workbook(excel_seguridad, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Helper para obtener año
    def get_year_label(val):
        if val is None:
            return None
        s = str(val).strip().lower()
        if 'tasa' in s:
            return None
        if s.endswith('.0'):
            s = s[:-2]
        match_yr = re.match(r'^(20\d{2})', s)
        if match_yr:
            return match_yr.group(1)
        return None

    # Detectar columnas de años dinámicamente en Fila 2 (index 1)
    col_inicio = {}
    if len(rows) > 1:
        for c_idx, val in enumerate(rows[1]):
            yr = get_year_label(val)
            if yr == '2023':
                col_inicio['2023'] = c_idx
            elif yr == '2024':
                col_inicio['2024'] = c_idx
            elif yr == '2025':
                col_inicio['2025'] = c_idx
            elif yr == '2026':
                col_inicio['2026*'] = c_idx
    print(f"  > Columnas de años detectadas: {col_inicio}")

    # Detectar columnas de metadatos en Fila 3 (index 2)
    col_ubicacion = None
    col_extension = None
    col_fecha = None
    if len(rows) > 2:
        for c_idx, val in enumerate(rows[2]):
            val_norm = normalize_text(val)
            if 'ubicacion' in val_norm:
                col_ubicacion = c_idx
            elif 'extension' in val_norm:
                col_extension = c_idx
            elif 'fecha' in val_norm:
                col_fecha = c_idx
    print(f"  > Columnas de metadatos detectadas: Ubicacion={col_ubicacion}, Extension={col_extension}, Fecha={col_fecha}")

    proyectos_list = []
    seguridad_list = []
    current_cat = ''

    # Almacenar nombres de proyectos para saber su UUID luego
    proyecto_uuid_map = {}

    for row in rows[4:]:
        if row[0] and isinstance(row[0], str) and row[0].strip():
            current_cat = row[0].strip()

        nombre = row[1]
        if not nombre or not isinstance(nombre, str) or not nombre.strip():
            continue
        nombre = nombre.strip()
        if nombre == 'Calle Benalzacar':
            nombre = 'Calle Benalcazar'

        # Insertar/Upsert Proyecto
        print(f"  > Proyecto encontrado: {nombre} ({current_cat})")
        proyecto_payload = {
            "nombre": nombre,
            "categoria": current_cat
        }
        if col_ubicacion is not None and col_ubicacion < len(row):
            proyecto_payload["ubicacion"] = str(row[col_ubicacion] or '').strip()
        if col_extension is not None and col_extension < len(row):
            proyecto_payload["extension"] = str(row[col_extension] or '').strip()
        if col_fecha is not None and col_fecha < len(row):
            proyecto_payload["fecha_inauguracion"] = str(row[col_fecha] or '').strip()
        
        try:
            res = supabase.table("proyectos").upsert(proyecto_payload, on_conflict="nombre").execute()
            if res.data:
                proj_id = res.data[0]['id']
                proyecto_uuid_map[normalize_text(nombre)] = proj_id
            else:
                print(f"Error al obtener ID del proyecto upserted: {nombre}")
                continue
        except Exception as e:
            print(f"Error al subir proyecto {nombre}: {e}")
            continue

        # Extraer estadísticas por año
        for anio, c_start in col_inicio.items():
            for i, var in enumerate(ALL_VARS):
                val = safe_float(row[c_start + i] if c_start + i < len(row) else None)
                if val is not None:
                    # '2026*' indica datos parciales (observados)
                    seguridad_list.append({
                        "proyecto_id": proj_id,
                        "anio": anio,
                        "variable": var,
                        "valor": val,
                        "tipo": "observado"
                    })

    if seguridad_list:
        print("  > Limpiando estadísticas de seguridad anteriores en la base de datos...")
        try:
            supabase.table("seguridad_estadisticas").delete().neq("id", -1).execute()
            print("  > Limpieza de seguridad completada.")
        except Exception as e:
            print(f"Error al limpiar seguridad: {e}")

        print(f"  > Subiendo {len(seguridad_list)} registros de estadísticas de seguridad...")
        # Subir en lotes de 1000
        for idx in range(0, len(seguridad_list), 1000):
            supabase.table("seguridad_estadisticas").insert(seguridad_list[idx:idx+1000]).execute()
        print("  > Carga de estadísticas completada.")

    # --- 2. PROCESAR Y CARGAR FICHAS DE SENDEROS (MATRIZ) ---
    print("\n[2/4] Procesando Fichas de Senderos desde senderos_matriz.py...")
    fichas_list = []
    for entrada in MATRIZ:
        nombre_proyecto = entrada['nombre']
        clave_resuelt = resolver_clave(nombre_proyecto)
        
        # Intentar vincular con el proyecto
        proyecto_id = None
        for nombre_normalizado, p_uuid in proyecto_uuid_map.items():
            if nombre_normalizado in normalize_text(nombre_proyecto) or normalize_text(nombre_proyecto) in nombre_normalizado:
                proyecto_id = p_uuid
                break
        
        fichas_list.append({
            "proyecto_id": proyecto_id,
            "anio": entrada['anio'],
            "ejecutor": entrada['ejecutor'],
            "presupuesto": entrada['presupuesto'],
            "extension": entrada['extension'],
            "beneficiarios": entrada['beneficiarios'],
            "fecha": entrada['fecha']
        })
    
    if fichas_list:
        print("  > Limpiando fichas de senderos anteriores...")
        try:
            supabase.table("senderos_fichas").delete().neq("id", -1).execute()
            print("  > Limpieza de fichas completada.")
        except Exception as e:
            print(f"Error al limpiar fichas: {e}")

        print(f"  > Subiendo {len(fichas_list)} fichas de senderos...")
        supabase.table("senderos_fichas").insert(fichas_list).execute()
        print("  > Carga de fichas de senderos completada.")

    # --- 3. PROCESAR Y CARGAR REGISTROS DE ECONOMÍA ---
    print("\n[3/4] Procesando Excels de Economía...")
    print("  > Limpiando registros de economía anteriores...")
    try:
        supabase.table("economia_registros").delete().neq("id", -1).execute()
        print("  > Limpieza de economía completada.")
    except Exception as e:
        print(f"Error al limpiar economía: {e}")

    economia_files = [
        DATA_DIR / 'resultado_cruce_predios_renovacion_v3.xlsx',
        DATA_DIR / 'resultado_cruce_predios_emision_v3.xlsx'
    ]

    SECTOR_HEADER_CANDIDATES = ('sector', 'proyecto', 'nombre proyecto', 'parroquia', 'ubicacion', 'barrio', 'zona')
    MOVIMIENTO_HEADER_CANDIDATES = ('tipo de movimiento(proceso', 'tipo de movimiento', 'proceso', 'movimiento', 'tipo movimiento')
    IMPRESION_HEADER_CANDIDATES = ('fecha de impresion', 'fecha de impresión', 'impresion', 'impresión', 'fecha impresion')

    for file_path in economia_files:
        if not file_path.is_file():
            print(f"  > Advertencia: No se encontró el archivo de economía {file_path.name}")
            continue

        print(f"  > Leyendo {file_path.name}...")
        wb_eco = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        ws_eco = wb_eco.active
        rows_eco = list(ws_eco.iter_rows(values_only=True))

        header_idx = find_header_row(rows_eco, SECTOR_HEADER_CANDIDATES + MOVIMIENTO_HEADER_CANDIDATES + IMPRESION_HEADER_CANDIDATES)
        if header_idx is None:
            header_idx = 0

        headers = list(rows_eco[header_idx]) if rows_eco else []
        sector_idx = match_column_name(headers, SECTOR_HEADER_CANDIDATES)
        movimiento_idx = match_column_name(headers, MOVIMIENTO_HEADER_CANDIDATES)
        fecha_idx = match_column_name(headers, IMPRESION_HEADER_CANDIDATES)

        if sector_idx is None and headers:
            sector_idx = 0

        economia_records = []
        for r in rows_eco[header_idx + 1:]:
            if not r:
                continue
            
            sector_raw = r[sector_idx] if sector_idx is not None and sector_idx < len(r) else None
            mov_raw = r[movimiento_idx] if movimiento_idx is not None and movimiento_idx < len(r) else None
            fecha_raw = r[fecha_idx] if fecha_idx is not None and fecha_idx < len(r) else None

            if sector_raw is None and mov_raw is None and fecha_raw is None:
                continue

            parsed_date = parse_flexible_date(fecha_raw)
            # En base de datos PostgreSQL, fecha_impresion es DATE, la enviamos como string YYYY-MM-DD o None
            date_str = parsed_date.strftime('%Y-%m-%d') if parsed_date else None

            economia_records.append({
                "sector_raw": str(sector_raw or '').strip(),
                "movimiento_raw": str(mov_raw or '').strip(),
                "fecha_impresion": date_str
            })

        if economia_records:
            print(f"  > Subiendo {len(economia_records)} registros de economía por lotes...")
            # Lotes de 1000
            import time
            for idx in range(0, len(economia_records), 1000):
                supabase.table("economia_registros").insert(economia_records[idx:idx+1000]).execute()
                time.sleep(0.5) # Pequeña pausa para evitar rate-limiting de Supabase
            print(f"  > Carga de {file_path.name} completada.")

    # --- 4. PROCESAR Y CARGAR VALOR DE SUELO ---
    print("\n[4/4] Procesando Excel de Valor de Suelo...")
    excel_suelo = None
    for pattern in ('*AIVAS*.xlsx', '*suelo*.xlsx', '*valor*.xlsx'):
        matches = sorted(DATA_DIR.glob(pattern))
        if matches:
            excel_suelo = matches[0]
            break

    VALOR_SUELO_HEADER_CANDIDATES = ('proyecto', 'descripcion', 'descripción', '2022-2023', '2024', '2026')
    PROYECTO_SUELO_HEADER_CANDIDATES = ('proyecto',)
    DESCRIPCION_SUELO_HEADER_CANDIDATES = ('descripcion', 'descripción')
    VALOR_SUELO_2022_HEADER_CANDIDATES = ('2022-2023', '2022 2023', '2022_2023')
    VALOR_SUELO_2024_HEADER_CANDIDATES = ('2024',)
    VALOR_SUELO_2026_HEADER_CANDIDATES = ('2026',)

    if excel_suelo and excel_suelo.is_file():
        print(f"  > Leyendo {excel_suelo.name}...")
        wb_suelo = openpyxl.load_workbook(excel_suelo, data_only=True, read_only=True)
        ws_suelo = wb_suelo.active
        rows_suelo = list(ws_suelo.iter_rows(values_only=True))

        header_row_idx = find_header_row(rows_suelo, VALOR_SUELO_HEADER_CANDIDATES, limit=30)
        if header_row_idx is None:
            header_row_idx = 0

        headers = list(rows_suelo[header_row_idx]) if rows_suelo else []
        proyecto_idx = match_column_name(headers, PROYECTO_SUELO_HEADER_CANDIDATES)
        descripcion_idx = match_column_name(headers, DESCRIPCION_SUELO_HEADER_CANDIDATES)
        valor_2022_idx = match_column_name(headers, VALOR_SUELO_2022_HEADER_CANDIDATES)
        valor_2024_idx = match_column_name(headers, VALOR_SUELO_2024_HEADER_CANDIDATES)
        valor_2026_idx = match_column_name(headers, VALOR_SUELO_2026_HEADER_CANDIDATES)

        suelo_records = []
        current_proyecto = ''
        
        for row in rows_suelo[header_row_idx + 1:]:
            if not row:
                continue

            proyecto_raw = row[proyecto_idx] if proyecto_idx is not None and proyecto_idx < len(row) else None
            if proyecto_raw is not None and str(proyecto_raw).strip():
                current_proyecto = str(proyecto_raw).strip()
            elif current_proyecto:
                proyecto_raw = current_proyecto

            descripcion_raw = row[descripcion_idx] if descripcion_idx is not None and descripcion_idx < len(row) else None
            v_2022 = safe_float(row[valor_2022_idx] if valor_2022_idx is not None and valor_2022_idx < len(row) else None)
            v_2024 = safe_float(row[valor_2024_idx] if valor_2024_idx is not None and valor_2024_idx < len(row) else None)
            v_2026 = safe_float(row[valor_2026_idx] if valor_2026_idx is not None and valor_2026_idx < len(row) else None)

            if not current_proyecto and not str(descripcion_raw or '').strip() and all(v is None for v in (v_2022, v_2024, v_2026)):
                continue

            suelo_records.append({
                "proyecto": current_proyecto,
                "descripcion": str(descripcion_raw or '').strip(),
                "valor_2022_2023": v_2022,
                "valor_2024": v_2024,
                "valor_2026": v_2026
            })

        if suelo_records:
            print("  > Limpiando registros de valor de suelo anteriores...")
            try:
                supabase.table("valor_suelo").delete().neq("id", -1).execute()
                print("  > Limpieza de valor de suelo completada.")
            except Exception as e:
                print(f"Error al limpiar valor de suelo: {e}")

            print(f"  > Subiendo {len(suelo_records)} registros de valor de suelo...")
            supabase.table("valor_suelo").insert(suelo_records).execute()
            print("  > Carga de valor de suelo completada.")
    else:
        print("  > Advertencia: No se encontró ningún archivo de Valor de Suelo en data/")

    print("\n=== MIGRACIÓN DE DATOS COMPLETADA CON ÉXITO ===")

if __name__ == '__main__':
    main()
