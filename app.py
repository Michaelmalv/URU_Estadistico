import html
import io
import hashlib
import os
import re
import textwrap
import unicodedata
from datetime import date, datetime
from pathlib import Path

import matplotlib.gridspec as gridspec
import matplotlib.patches as mpatches
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
import streamlit as st

from senderos_imagenes import imagenes_sendero
from senderos_matriz import es_sendero_seguro, fichas_sendero


INCIDENTES = [
    'Daño a propiedad pública y privada',
    'Escándalos',
    'Eventos clandestinos',
    'Libadores',
    'Venta y consumo de sustancias',
]
DELITOS = [
    'Robo a carros', 'Robo a motos', 'Robo a personas',
    'Robo a unidades económicas', 'Robo de autopartes', 'Robo a domicilios',
]
ALL_VARS = INCIDENTES + DELITOS

COL_INICIO = {'2023': 5, '2024': 16, '2025': 27, '2026*': 38}
COL_TASA_EXCEL = {'2023-2024': 49, '2024-2025': 60}
PERIODOS = list(COL_INICIO.keys())
MESES_2026 = 4


def _anio_base(periodo: str) -> int:
    match = re.search(r'\d{4}', periodo)
    return int(match.group()) if match else 0


def _periodos_actuales_disponibles(anios_anterior):
    if not anios_anterior:
        return PERIODOS
    max_anterior = max(_anio_base(anio) for anio in anios_anterior)
    return [anio for anio in PERIODOS if _anio_base(anio) > max_anterior]


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / 'data'
DEFAULT_EXCEL_NAMES = (
    'Evaluación de proyectos estratégicos_ SEGURIDAD.xlsx',
    'datos_seguridad.xlsx',
    'datos.xlsx',
)
DEFAULT_ECONOMIA_EXCEL_NAMES = (
    'resultado_cruce_predios_renovacion_v3.xlsx',
    'resultado_cruce_predios_emision_v3.xlsx',
)
DEFAULT_VALOR_SUELO_EXCEL_NAMES = (
    'AIVAS_cruce_codigos_nuevo.xlsx',
    'valor_de_suelo.xlsx',
    'valor_suelo.xlsx',
    'valores_suelo.xlsx',
)

MONTHS_ES = {
    'enero': 1,
    'febrero': 2,
    'marzo': 3,
    'abril': 4,
    'mayo': 5,
    'junio': 6,
    'julio': 7,
    'agosto': 8,
    'septiembre': 9,
    'setiembre': 9,
    'octubre': 10,
    'noviembre': 11,
    'diciembre': 12,
}

SECTOR_HEADER_CANDIDATES = (
    'sector', 'proyecto', 'nombre proyecto', 'nombre del proyecto',
    'parroquia', 'ubicacion', 'ubicación', 'barrio', 'zona'
)
MOVIMIENTO_HEADER_CANDIDATES = (
    'tipo de movimiento(proceso', 'tipo de movimiento (proceso',
    'tipo de movimiento', 'proceso', 'movimiento', 'tipo movimiento'
)
IMPRESION_HEADER_CANDIDATES = (
    'fecha de impresion', 'fecha de impresión', 'impresion', 'impresión',
    'fecha impresion', 'fecha impresión'
)
VALOR_SUELO_HEADER_CANDIDATES = (
    'proyecto', 'descripcion', 'descripción', '2022-2023', '2024', '2026'
)
PROYECTO_SUELO_HEADER_CANDIDATES = ('proyecto',)
DESCRIPCION_SUELO_HEADER_CANDIDATES = ('descripcion', 'descripción')
VALOR_SUELO_2022_HEADER_CANDIDATES = ('2022-2023', '2022 2023', '2022_2023')
VALOR_SUELO_2024_HEADER_CANDIDATES = ('2024',)
VALOR_SUELO_2026_HEADER_CANDIDATES = ('2026',)

VALOR_SUELO_CATEGORIAS = {
    'Zonas Metro': [
        'AIVAS EL EJIDO',
        'AIVAS ALAMEDA',
        'AIVAS CARDENAL DE LA TORRE',
        'AIVAS EL LABRADOR',
        'AIVAS EL RECREO',
        'AIVAS IÑAQUITO',
        'AIVAS JIPIJAPA',
        'AIVAS LA CAROLINA',
        'AIVAS LA MAGDALENA',
        'AIVAS LA PRADERA',
        'AIVAS MORAN VALVERDE',
        'AIVAS QUITUMBE',
        'AIVAS SAN FRANCISCO',
        'AIVAS SOLANDA',
        'AIVAS UNIVERSIDAD CENTRAL',
    ],
    'Rehabilitación del Espacio Público y Centro Histórico': [
        'AIVAS PARQUE NAVARRO',
        'AIVAS BENALCAZAR',
        'AIVAS ROCAFUERTE',
        'AIVAS TRIBUNA DE LOS SHYRIS',
    ],
    'Senderos Seguros': [
        'AIVAS AV. AJAVÍ',
        'AIVAS LA ECUATORIANA',
        'AIVAS CALLE RUIZ DE CASTILLA',
        'AIVAS CALLE RÍO DE JANEIRO',
        'AIVAS COMITÉ DEL PUEBLO',
        'AIVAS LA MARISCAL',
        'AIVAS RAMÓN BORJA',
        'AIVAS AV. 2 DE AGOSTO',
        'AIVAS AV. CARAPUNGO',
        'AIVAS COLINAS DEL NORTE',
        'AIVAS CALLE CALDAS Y ANTEPARA',
        'AIVAS CALLE JUAN MONTALVO',
        'AIVAS CALLE GABRIEL GARCÍA MORENO',
        'AIVAS CALLE LIZARDO RUIZ',
        'AIVAS VÍAS DEL FERROCARRIL',
        'AIVAS AV. CACHA',
        'AIVAS CONOCOTO',
        'AIVAS NANEGALITO',
        'AIVAS PATRIA',
        'AIVAS AV. MICHELENA',
        'AIVAS CALLE LUIS LÓPEZ',
        'AIVAS AV. COLÓN',
        'AIVAS ISLA TORTUGA',
    ],
}
VALOR_SUELO_CATEGORIA_ORDER = [
    'Todas',
    'Zonas Metro',
    'Rehabilitación del Espacio Público y Centro Histórico',
    'Senderos Seguros',
]

PORTAL_TITLE = 'PORTAL DE EVALUACIÓN DE PROYECTOS ESTRATÉGICOS'
PORTAL_SUBTITLE = 'DIRECCIÓN DE DESARROLLO URBANÍSTICO'
HEADER_IMAGE_CANDIDATES = (
    DATA_DIR / 'header.png',
    DATA_DIR / 'header.jpg',
    DATA_DIR / 'header.jpeg',
    DATA_DIR / 'header.webp',
)


def render_portal_header():
    for image_path in HEADER_IMAGE_CANDIDATES:
        if image_path.is_file():
            st.image(str(image_path), use_container_width=True)
            st.markdown(
                """
                <div style="margin-top:0.6rem; padding:0.4rem 0 0.8rem 0;">
                  <h1 style="margin:0; font-size:32px; line-height:1.05; letter-spacing:0.02em; color:#24367f; text-transform:uppercase;">PORTAL DE EVALUACIÓN DE PROYECTOS ESTRATÉGICOS</h1>
                  <p style="margin:6px 0 0; font-size:12px; color:#56607c; text-transform:uppercase; font-weight:700;">DIRECCIÓN DE DESARROLLO URBANÍSTICO</p>
                </div>
                """,
                unsafe_allow_html=True,
            )
            return
    st.warning('Falta la imagen del header. Coloca el archivo en data/header.png para mostrarla en el inicio.')


def find_default_excel():
    """Busca el Excel local: variable de entorno, nombres por defecto o cualquier .xlsx en data/."""
    env_path = os.environ.get('SEGURIDAD_EXCEL')
    if env_path:
        path = Path(env_path).expanduser().resolve()
        if path.is_file():
            return path
    for name in DEFAULT_EXCEL_NAMES:
        path = DATA_DIR / name
        if path.is_file():
            return path
    if DATA_DIR.is_dir():
        files = sorted(DATA_DIR.glob('*.xlsx'))
        if files:
            return files[0]
    return None


@st.cache_data(show_spinner='Cargando datos del Excel…')
def load_workbook_data(cache_key: str, file_bytes: bytes):
    return parse_workbook(file_bytes)


@st.cache_data(show_spinner='Cargando datos económicos…')
def load_economia_data(cache_key: str, file_bytes: bytes, source_label: str):
    return parse_economia_workbook(file_bytes, source_label)


@st.cache_data(show_spinner='Cargando datos de valor de suelo…')
def load_valor_suelo_data(cache_key: str, file_bytes: bytes, source_label: str):
    return parse_valor_suelo_workbook(file_bytes, source_label)


def workbook_cache_key(source_label: str, file_bytes: bytes, file_path: Path | None = None):
    if file_path is not None:
        stat = file_path.stat()
        return f'file:{file_path}:{stat.st_mtime_ns}:{stat.st_size}'
    digest = hashlib.md5(file_bytes).hexdigest()
    return f'upload:{source_label}:{digest}'


def safe_float(v):
    if v is None or str(v).strip() in ('', '-'):
        return None
    try:
        return float(v)
    except:
        return None


def normalize_text(value):
    if value is None:
        return ''
    text = str(value).strip().lower()
    text = unicodedata.normalize('NFKD', text).encode('ascii', 'ignore').decode('ascii')
    text = re.sub(r'[^a-z0-9]+', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    return text


def parse_flexible_date(value):
    if value is None:
        return None
    if isinstance(value, pd.Timestamp):
        value = value.to_pydatetime()
    if isinstance(value, datetime):
        return {
            'date': value.date(),
            'year': value.year,
            'month': value.month,
            'day': value.day,
            'precision': 'date',
            'raw': value,
        }
    if isinstance(value, date):
        return {
            'date': value,
            'year': value.year,
            'month': value.month,
            'day': value.day,
            'precision': 'date',
            'raw': value,
        }
    if isinstance(value, (int, float)) and not pd.isna(value):
        year = int(value)
        if 1900 <= year <= 2100 and abs(float(value) - year) < 1e-6:
            return {
                'date': date(year, 1, 1),
                'year': year,
                'month': None,
                'day': None,
                'precision': 'year',
                'raw': value,
            }
        return None

    text = str(value).strip()
    if not text or normalize_text(text) in ('no', 'nan', 'none', '-', '—'):
        return None

    dt = pd.to_datetime(text, dayfirst=True, errors='coerce')
    if pd.notna(dt):
        return {
            'date': dt.date(),
            'year': dt.year,
            'month': dt.month,
            'day': dt.day,
            'precision': 'date',
            'raw': text,
        }

    norm = normalize_text(text)
    m = re.search(r'(?P<day>\d{1,2})\s*(?:de\s*)?(?P<month>[a-zñ]+)\s*(?:de\s*)?(?P<year>\d{4})', norm)
    if m:
        month = MONTHS_ES.get(m.group('month'))
        if month:
            year = int(m.group('year'))
            day = int(m.group('day'))
            try:
                dt = date(year, month, day)
                return {
                    'date': dt,
                    'year': year,
                    'month': month,
                    'day': day,
                    'precision': 'date',
                    'raw': text,
                }
            except ValueError:
                pass

    m = re.search(r'(?P<month>[a-zñ]+)\s+(?P<year>\d{4})', norm)
    if m:
        month = MONTHS_ES.get(m.group('month'))
        if month:
            year = int(m.group('year'))
            try:
                dt = date(year, month, 1)
                return {
                    'date': dt,
                    'year': year,
                    'month': month,
                    'day': None,
                    'precision': 'month',
                    'raw': text,
                }
            except ValueError:
                pass

    m = re.search(r'\b(19|20)\d{2}\b', norm)
    if m:
        year = int(m.group(0))
        return {
            'date': date(year, 1, 1),
            'year': year,
            'month': None,
            'day': None,
            'precision': 'year',
            'raw': text,
        }

    return None


def match_column_name(headers, candidates):
    normalized_headers = [normalize_text(h) for h in headers]
    normalized_candidates = [normalize_text(c) for c in candidates]
    for candidate in normalized_candidates:
        for idx, header in enumerate(normalized_headers):
            if header == candidate or candidate in header or header in candidate:
                return idx
    return None


def find_header_row(rows, candidates, limit=20):
    for idx, row in enumerate(rows[:limit]):
        normalized = ' | '.join(normalize_text(cell) for cell in row if cell is not None)
        if any(normalize_text(candidate) in normalized for candidate in candidates):
            return idx
    return None


def parse_economia_workbook(file_bytes, source_label):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header_row_idx = find_header_row(rows, SECTOR_HEADER_CANDIDATES + MOVIMIENTO_HEADER_CANDIDATES + IMPRESION_HEADER_CANDIDATES)
    if header_row_idx is None:
        header_row_idx = 0
    headers = list(rows[header_row_idx]) if rows else []
    sector_idx = match_column_name(headers, SECTOR_HEADER_CANDIDATES)
    movimiento_idx = match_column_name(headers, MOVIMIENTO_HEADER_CANDIDATES)
    fecha_idx = match_column_name(headers, IMPRESION_HEADER_CANDIDATES)
    if sector_idx is None and headers:
        sector_idx = 0

    records = []
    for row in rows[header_row_idx + 1:]:
        if not row:
            continue
        sector_raw = row[sector_idx] if sector_idx is not None and sector_idx < len(row) else None
        movimiento_raw = row[movimiento_idx] if movimiento_idx is not None and movimiento_idx < len(row) else None
        fecha_raw = row[fecha_idx] if fecha_idx is not None and fecha_idx < len(row) else None

        if sector_raw is None and movimiento_raw is None and fecha_raw is None:
            continue

        records.append({
            'sector_raw': str(sector_raw or '').strip(),
            'movimiento_raw': str(movimiento_raw or '').strip(),
            'fecha_impresion_raw': fecha_raw,
            'fecha_impresion': parse_flexible_date(fecha_raw),
            'source_label': source_label,
        })

    return {
        'records': records,
        'headers': headers,
        'sector_idx': sector_idx,
        'movimiento_idx': movimiento_idx,
        'fecha_idx': fecha_idx,
        'header_row_idx': header_row_idx,
    }


def parse_valor_suelo_workbook(file_bytes, source_label):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header_row_idx = find_header_row(rows, VALOR_SUELO_HEADER_CANDIDATES, limit=30)
    if header_row_idx is None:
        header_row_idx = 0

    headers = list(rows[header_row_idx]) if rows else []
    proyecto_idx = match_column_name(headers, PROYECTO_SUELO_HEADER_CANDIDATES)
    descripcion_idx = match_column_name(headers, DESCRIPCION_SUELO_HEADER_CANDIDATES)
    valor_2022_idx = match_column_name(headers, VALOR_SUELO_2022_HEADER_CANDIDATES)
    valor_2024_idx = match_column_name(headers, VALOR_SUELO_2024_HEADER_CANDIDATES)
    valor_2026_idx = match_column_name(headers, VALOR_SUELO_2026_HEADER_CANDIDATES)

    records = []
    current_proyecto = ''
    current_proyecto_key = ''

    for row in rows[header_row_idx + 1:]:
        if not row:
            continue

        proyecto_raw = row[proyecto_idx] if proyecto_idx is not None and proyecto_idx < len(row) else None
        if proyecto_raw is not None and str(proyecto_raw).strip():
            current_proyecto = str(proyecto_raw).strip()
            current_proyecto_key = normalize_text(current_proyecto)
        elif current_proyecto:
            proyecto_raw = current_proyecto

        descripcion_raw = row[descripcion_idx] if descripcion_idx is not None and descripcion_idx < len(row) else None
        valor_2022 = safe_float(row[valor_2022_idx] if valor_2022_idx is not None and valor_2022_idx < len(row) else None)
        valor_2024 = safe_float(row[valor_2024_idx] if valor_2024_idx is not None and valor_2024_idx < len(row) else None)
        valor_2026 = safe_float(row[valor_2026_idx] if valor_2026_idx is not None and valor_2026_idx < len(row) else None)

        if not current_proyecto_key and not str(descripcion_raw or '').strip() and all(v is None for v in (valor_2022, valor_2024, valor_2026)):
            continue

        records.append({
            'proyecto_raw': current_proyecto,
            'proyecto_key': current_proyecto_key,
            'descripcion_raw': str(descripcion_raw or '').strip(),
            'valor_2022_2023': valor_2022,
            'valor_2024': valor_2024,
            'valor_2026': valor_2026,
            'source_label': source_label,
        })

    if not records:
        return {
            'records': [],
            'resumen': pd.DataFrame(),
            'headers': headers,
            'header_row_idx': header_row_idx,
            'source_label': source_label,
        }

    df = pd.DataFrame(records)
    df = df[df['proyecto_key'].astype(str).str.strip() != '']
    if df.empty:
        return {
            'records': records,
            'resumen': pd.DataFrame(),
            'headers': headers,
            'header_row_idx': header_row_idx,
            'source_label': source_label,
        }

    resumen = (
        df.groupby('proyecto_key', dropna=False)
        .agg(
            proyecto=('proyecto_raw', 'first'),
            sectores=('descripcion_raw', lambda s: int(s.astype(str).str.strip().ne('').sum())),
            descripciones=('descripcion_raw', lambda s: sorted({x for x in (str(v).strip() for v in s) if x})),
            valor_2022_2023=('valor_2022_2023', 'mean'),
            valor_2024=('valor_2024', 'mean'),
            valor_2026=('valor_2026', 'mean'),
        )
        .reset_index(drop=True)
    )
    resumen['descripciones'] = resumen['descripciones'].apply(lambda values: ', '.join(values) if values else '')
    resumen['proyecto'] = resumen['proyecto'].fillna('').astype(str).str.strip()
    resumen = resumen.sort_values('proyecto', kind='stable').reset_index(drop=True)

    return {
        'records': records,
        'resumen': resumen,
        'headers': headers,
        'header_row_idx': header_row_idx,
        'source_label': source_label,
    }


def build_sector_alias_map(proyectos):
    alias_map = {}
    for nombre, proyecto in proyectos.items():
        alias_map[normalize_text(nombre)] = nombre
        if proyecto.get('ubicacion'):
            alias_map[normalize_text(proyecto['ubicacion'])] = nombre
    return alias_map


def resolve_sector_name(raw_sector, alias_map):
    normalized = normalize_text(raw_sector)
    if not normalized:
        return None
    if normalized in alias_map:
        return alias_map[normalized]
    for alias, canonical in alias_map.items():
        if alias and (alias in normalized or normalized in alias):
            return canonical
    return None


def classify_business(record, fecha_referencia):
    movimiento = normalize_text(record.get('movimiento_raw'))
    fecha_impresion = record.get('fecha_impresion')

    if 'renov' in movimiento:
        return 'renovado', 'proceso'
    if 'emisi' in movimiento:
        return 'abierto', 'proceso'

    if fecha_referencia and fecha_impresion:
        ref_precision = fecha_referencia.get('precision')
        imp_precision = fecha_impresion.get('precision')
        if ref_precision == 'date' and imp_precision == 'date':
            if fecha_impresion['date'] < fecha_referencia['date']:
                return 'renovado', 'fecha'
            if fecha_impresion['date'] >= fecha_referencia['date']:
                return 'abierto', 'fecha'
        else:
            if fecha_impresion['year'] < fecha_referencia['year']:
                return 'renovado', 'anio'
            if fecha_impresion['year'] > fecha_referencia['year']:
                return 'abierto', 'anio'

    return 'indeterminado', 'sin_datos'


def is_after_or_equal_reference(fecha_impresion, fecha_referencia):
    if not fecha_impresion or not fecha_referencia:
        return True

    ref_precision = fecha_referencia.get('precision')
    imp_precision = fecha_impresion.get('precision')

    if ref_precision == 'date' and imp_precision == 'date':
        return fecha_impresion['date'] >= fecha_referencia['date']

    return fecha_impresion['year'] >= fecha_referencia['year']


def find_default_economia_excels():
    env_renovacion = os.environ.get('ECONOMIA_RENOVACION_EXCEL')
    env_emision = os.environ.get('ECONOMIA_EMISION_EXCEL')
    encontrados = []

    for env_path in (env_renovacion, env_emision):
        if env_path:
            path = Path(env_path).expanduser().resolve()
            if path.is_file():
                encontrados.append(path)

    for name in DEFAULT_ECONOMIA_EXCEL_NAMES:
        path = DATA_DIR / name
        if path.is_file() and path not in encontrados:
            encontrados.append(path)

    if DATA_DIR.is_dir():
        for path in sorted(DATA_DIR.glob('resultado_cruce_predios_*v3.xlsx')):
            if path not in encontrados:
                encontrados.append(path)

    return encontrados[:2]


def find_default_valor_suelo_excel():
    env_path = os.environ.get('VALOR_SUELO_EXCEL')
    if env_path:
        path = Path(env_path).expanduser().resolve()
        if path.is_file():
            return path

    for name in DEFAULT_VALOR_SUELO_EXCEL_NAMES:
        path = DATA_DIR / name
        if path.is_file():
            return path

    if DATA_DIR.is_dir():
        for pattern in ('*AIVAS*.xlsx', '*suelo*.xlsx', '*valor*.xlsx'):
            matches = sorted(DATA_DIR.glob(pattern))
            if matches:
                return matches[0]

    return None


def resolve_valor_suelo_categoria(proyecto: str):
    normalizado = normalize_text(proyecto)
    for categoria, proyectos_categoria in VALOR_SUELO_CATEGORIAS.items():
        for candidato in proyectos_categoria:
            if normalizado == normalize_text(candidato):
                return categoria
    return 'Sin categoría'


def filtrar_valor_suelo_por_categoria(resumen_suelo: pd.DataFrame, categoria: str):
    if categoria == 'Todas':
        return resumen_suelo.copy()
    if resumen_suelo.empty:
        return resumen_suelo.copy()
    filtrado = resumen_suelo[resumen_suelo['categoria'] == categoria].copy()
    return filtrado.reset_index(drop=True)


def parse_workbook(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    categorias = {}
    proyectos = {}
    current_cat = ''

    for row in rows[4:]:
        if row[0] and isinstance(row[0], str) and row[0].strip():
            current_cat = row[0].strip()
            if current_cat not in categorias:
                categorias[current_cat] = []
        nombre = row[1]
        if not nombre or not isinstance(nombre, str):
            continue
        nombre = nombre.strip()

        anios = {}
        for anio, c in COL_INICIO.items():
            anios[anio] = {var: safe_float(row[c + i] if c + i < len(row) else None)
                           for i, var in enumerate(ALL_VARS)}

        tasas_excel = {}
        for llave, c in COL_TASA_EXCEL.items():
            tasas_excel[llave] = {var: safe_float(row[c + i] if c + i < len(row) else None)
                                  for i, var in enumerate(ALL_VARS)}

        proyeccion_2026 = {
            var: (anios['2026*'][var] * 12 / MESES_2026
                  if anios['2026*'][var] is not None else None)
            for var in ALL_VARS
        }

        tiene_datos = any(any(v is not None for v in d.values()) for d in anios.values())
        tiene_metadatos = any(str(row[i] or '').strip() for i in (2, 3, 4))
        if tiene_datos or tiene_metadatos:
            proyectos[nombre] = {
                'tiene_estadisticas': tiene_datos,
                'categoria': current_cat,
                'programa': current_cat,
                'ubicacion': str(row[2] or '').strip(),
                'extension': str(row[3] or '').strip(),
                'fecha': str(row[4] or '').strip(),
                'anios': anios,
                'tasas_excel': tasas_excel,
                'proyeccion_2026': proyeccion_2026
            }
            if current_cat:
                categorias[current_cat].append(nombre)

    return categorias, proyectos


def tasa_calc(a, b):
    if a is None or b is None or a == 0:
        return None
    return (b - a) / a


def generar_tabla(p, anios_ant, anios_act):
    """Genera tabla comparativa con incidentes y delitos."""
    todos = anios_ant + anios_act
    n = len(todos)
    anio_ref_ant = anios_ant[-1]
    anio_ref_act = anios_act[0]
    llave = f'{anio_ref_ant}-{anio_ref_act}'
    tasas_pre = p.get('tasas_excel', {}).get(llave, {})

    C_HDR='#1a1a2e'; C_ANT='#2c3e50'; C_ACT='#1a5276'; C_TASA='#4a235a'
    RODD='#f7f7f7'; REVEN='#ffffff'; TBAG='#f0e6f6'
    escala = 1.55
    ROW_H = 0.58 * escala
    HDR_H = 0.82 * escala
    GAP_H = 0.55 * escala
    TTL_H = 1.15 * escala
    FT_H = 0.40 * escala
    h_inc = HDR_H + len(INCIDENTES) * ROW_H
    h_del = HDR_H + len(DELITOS) * ROW_H
    fig_w = min(max(16., 13. + n * 1.6), 28.)
    fig_h = (TTL_H + h_inc + GAP_H + h_del + FT_H) * 1.05

    fig = plt.figure(figsize=(fig_w, fig_h), facecolor='white', dpi=120)
    gs = gridspec.GridSpec(5, 1, figure=fig,
         height_ratios=[TTL_H, h_inc, GAP_H, h_del, FT_H], hspace=0)

    ax0 = fig.add_subplot(gs[0]); ax0.set_facecolor(C_HDR); ax0.axis('off')
    ax0.text(.5, .68, p['programa'].upper(), transform=ax0.transAxes,
             ha='center', va='center', fontsize=18, fontweight='bold', color='white')
    ax0.text(.5, .22,
             f"{p['ubicacion']} | {p['extension']} | Inauguración: {p['fecha']}",
             transform=ax0.transAxes, ha='center', va='center', fontsize=10, color='#bbb')

    def draw_table(ax, filas_vars, tipo_hdr):
        ax.set_facecolor('white'); ax.axis('off')
        n_anios = len(todos)
        n = len(filas_vars); LW = .13; TX = LW + .005; TW = 1. - TX
        ht = HDR_H + n * ROW_H; HF = HDR_H / ht; RF = ROW_H / ht
        WN = TW * .40; WA = TW * (.45 / n_anios); WT = TW * .15
        cxs = [TX] + [TX + WN + k * WA for k in range(n_anios)] + [TX + WN + n_anios * WA]
        cws = [WN] + [WA] * n_anios + [WT]
        ax.text(LW * .45, .5, tipo_hdr, transform=ax.transAxes,
                ha='center', va='center', fontsize=11, fontweight='bold', color='#1a1a2e')
        hdrs = [tipo_hdr] + todos + [f'Tasa\n{anio_ref_ant}→{anio_ref_act}']
        hcols = ([C_ANT] + [C_ANT] * len(anios_ant) + [C_ACT] * len(anios_act) + [C_TASA])
        for lbl, cx, cw, bg in zip(hdrs, cxs, cws, hcols):
            ax.add_patch(mpatches.FancyBboxPatch((cx, 1 - HF), cw, HF,
                boxstyle='square,pad=0', facecolor=bg, edgecolor='white', lw=.8,
                transform=ax.transAxes, clip_on=False))
            ax.text(cx + cw / 2, 1 - HF / 2, lbl, transform=ax.transAxes,
                    ha='center', va='center', fontsize=11, fontweight='bold',
                    color='white', linespacing=1.3)
        for i, var in enumerate(filas_vars):
            yt = 1 - HF - i * RF; yb = yt - RF
            bg = RODD if i % 2 == 0 else REVEN
            va = p['anios'][anio_ref_ant].get(var)
            vb = p['anios'][anio_ref_act].get(var)
            t = tasas_pre.get(var) if tasas_pre.get(var) is not None else tasa_calc(va, vb)
            ax.add_patch(mpatches.FancyBboxPatch((cxs[0], yb), cws[0], RF,
                boxstyle='square,pad=0', facecolor=bg, edgecolor='white', lw=.4,
                transform=ax.transAxes, clip_on=False))
            ax.text(cxs[0] + .007, (yt + yb) / 2, var, transform=ax.transAxes,
                    ha='left', va='center', fontsize=11, color='#333')
            for k, anio in enumerate(todos):
                val = p['anios'][anio].get(var); cx = cxs[1 + k]; cw = cws[1 + k]
                fd = bg if anio in anios_ant else '#eaf4fb'
                ax.add_patch(mpatches.FancyBboxPatch((cx, yb), cw, RF,
                    boxstyle='square,pad=0', facecolor=fd, edgecolor='white', lw=.4,
                    transform=ax.transAxes, clip_on=False))
                ax.text(cx + cw / 2, (yt + yb) / 2, str(int(val)) if val else '-',
                        transform=ax.transAxes, ha='center', va='center',
                        fontsize=12, color='#333')
            ct = cxs[-1]; cw = cws[-1]
            ax.add_patch(mpatches.FancyBboxPatch((ct, yb), cw, RF,
                boxstyle='square,pad=0', facecolor=TBAG, edgecolor='white', lw=.4,
                transform=ax.transAxes, clip_on=False))
            fmt_t = f'{t*100:+.0f}%' if t is not None else 'N/A'
            col_t = '#c0392b' if t and t > 0 else '#27ae60' if t and t < 0 else '#777'
            ax.text(ct + cw / 2, (yt + yb) / 2, fmt_t,
                    transform=ax.transAxes, ha='center', va='center',
                    fontsize=12, fontweight='bold', color=col_t)

    ax1 = fig.add_subplot(gs[1]); draw_table(ax1, INCIDENTES, 'INCIDENTE')
    ax2 = fig.add_subplot(gs[2]); ax2.axis('off'); ax2.set_facecolor('white')
    ax3 = fig.add_subplot(gs[3]); draw_table(ax3, DELITOS, 'DELITO')
    ax4 = fig.add_subplot(gs[4]); ax4.axis('off'); ax4.set_facecolor('white')
    ax4.text(.01, .5, f'Fuente: ECU 911 · PPNN | Tasa: {anio_ref_ant}→{anio_ref_act}',
             transform=ax4.transAxes, ha='left', va='center', fontsize=9, color='#777')
    fig.subplots_adjust(left=.02, right=.98, top=.98, bottom=.02, hspace=0)
    return fig


def mostrar_figura_alta_res(fig, dpi_pantalla: int = 220):
    """Muestra la figura a ancho completo en alta resolución (evita pixelación al ampliar)."""
    buf = io.BytesIO()
    fig.savefig(
        buf, format='png', dpi=dpi_pantalla, bbox_inches='tight',
        facecolor='white', pad_inches=0.2,
    )
    buf.seek(0)
    st.image(buf, use_container_width=True)


def _etiqueta_grafico(nombre: str, ancho: int = 16) -> str:
    """Parte el nombre en líneas para evitar solapamiento en el eje X."""
    lineas = textwrap.wrap(nombre, width=ancho, break_long_words=False)
    return '\n'.join(lineas[:3]) if lineas else nombre


def generar_grafico_resumen(p, anios_ant, anios_act):
    todos = anios_ant + anios_act
    n = len(todos)
    tiene_2026 = '2026*' in anios_act
    anio_ref_ant = anios_ant[-1]
    anio_ref_act = anios_act[0]

    COLS_ANT = ['#1a5276', '#2e86c1', '#5dade2']
    COLS_ACT = ['#1e8449', '#27ae60', '#82e0aa']
    colores = ([COLS_ANT[i % 3] for i in range(len(anios_ant))] +
               [COLS_ACT[i % 3] for i in range(len(anios_act))])

    max_cats = max(len(INCIDENTES), len(DELITOS))
    fig_w = max(14., 10. + max_cats * 2.2 + n * 1.0)
    alto_panel = 6.5 + max_cats * 0.35
    fig_h = alto_panel * 2 + 1.8
    fig, axes = plt.subplots(2, 1, figsize=(fig_w, fig_h), dpi=120)
    fig.patch.set_facecolor('white')
    fig.suptitle('Resumen comparativo', fontsize=18, fontweight='bold', y=0.995)

    def _grupo(ax, vlist, titulo):
        x = np.arange(len(vlist))
        w = 0.75 / n
        max_val = 0
        for k, anio in enumerate(todos):
            vals = [p['anios'][anio].get(v) or 0 for v in vlist]
            max_val = max(max_val, max(vals))
            offset = (k - n / 2 + .5) * w
            bars = ax.bar(x + offset, vals, width=w * .9, color=colores[k], label=anio)
            for bar, val in zip(bars, vals):
                if val > 0:
                    ax.text(
                        bar.get_x() + bar.get_width() / 2,
                        bar.get_height() + max_val * 0.02,
                        str(int(val)), ha='center', va='bottom', fontsize=11,
                    )

        if tiene_2026:
            k26 = todos.index('2026*')
            offset26 = (k26 - n / 2 + .5) * w
            for j, var in enumerate(vlist):
                real = p['anios']['2026*'].get(var) or 0
                proy = p['proyeccion_2026'].get(var) or 0
                if proy > real:
                    ax.bar(
                        x[j] + offset26, proy - real, width=w * .9,
                        bottom=real, color='none', edgecolor='#27ae60', lw=1.2, hatch='//',
                    )
                    ax.text(
                        x[j] + offset26, proy + max_val * 0.04, f'~{int(proy)}',
                        ha='center', va='bottom', fontsize=10, color='#1e8449',
                    )

        techo = max(max_val * 1.45, 1)
        for j, var in enumerate(vlist):
            va_v = p['anios'][anio_ref_ant].get(var)
            vb_v = p['anios'][anio_ref_act].get(var)
            t = tasa_calc(va_v, vb_v)
            if t is None:
                continue
            flecha = '↑' if t > 0 else '↓'
            color = '#c0392b' if t > 0 else '#27ae60'
            ax.text(
                x[j], techo * 0.96, f'{flecha}{abs(t * 100):.0f}%',
                ha='center', va='top', fontsize=12, fontweight='bold', color=color,
            )

        ancho_etiq = 16 if len(vlist) >= 6 else 20
        etiq = [_etiqueta_grafico(v, ancho_etiq) for v in vlist]
        ax.set_xticks(x)
        ax.set_xticklabels(etiq, fontsize=11, rotation=30, ha='right')
        ax.set_title(titulo, fontsize=14, backgroundcolor='#2c3e50', color='white', pad=12)
        ax.set_ylim(0, techo)
        ax.legend(fontsize=11, loc='upper left', framealpha=0.9)
        ax.margins(x=0.06)
        ax.tick_params(axis='x', pad=10)
        ax.tick_params(axis='y', labelsize=11)

    _grupo(axes[0], INCIDENTES, 'INCIDENTES (ECU 911)')
    _grupo(axes[1], DELITOS, 'DELITOS (PPNN)')
    fig.tight_layout(rect=[0, 0.02, 1, 0.97], h_pad=4.0)
    fig.subplots_adjust(hspace=0.45)
    return fig


def export_csv(proyectos):
    registros = []
    for nombre, p in proyectos.items():
        for anio, datos in p['anios'].items():
            for var, val in datos.items():
                if val is not None:
                    tipo = 'Incidente' if var in INCIDENTES else 'Delito'
                    registros.append({
                        'Categoria': p['categoria'],
                        'Proyecto': nombre,
                        'Ubicacion': p['ubicacion'],
                        'Extension': p['extension'],
                        'Fecha': p['fecha'],
                        'Anio': anio,
                        'Tipo': tipo,
                        'Variable': var,
                        'Valor': int(round(val))
                    })
    df = pd.DataFrame(registros)
    # proyeccion
    for nombre, p in proyectos.items():
        for var, val in p['proyeccion_2026'].items():
            if val is not None:
                tipo = 'Incidente' if var in INCIDENTES else 'Delito'
                df = df.append({
                    'Categoria': p['categoria'],
                    'Proyecto': nombre,
                    'Ubicacion': p['ubicacion'],
                    'Extension': p['extension'],
                    'Fecha': p['fecha'],
                    'Anio': '2026_proyeccion',
                    'Tipo': tipo,
                    'Variable': var,
                    'Valor': int(round(val))
                }, ignore_index=True)
    return df


def _ficha_css():
    tema_base = str(st.get_option('theme.base') or '').lower()
    if tema_base == 'dark':
        label_color = '#d2deec'
        value_color = '#ffffff'
        title_color = '#ffffff'
        border_color = 'rgba(255, 255, 255, 0.18)'
        card_bg = 'rgba(15, 23, 42, 0.90)'
        title_bg = 'rgba(59, 130, 246, 0.14)'
        title_border = 'rgba(148, 163, 184, 0.30)'
    else:
        label_color = '#334155'
        value_color = '#0f172a'
        title_color = '#0f172a'
        border_color = 'rgba(148, 163, 184, 0.55)'
        card_bg = 'rgba(248, 250, 252, 0.92)'
        title_bg = 'rgba(59, 130, 246, 0.06)'
        title_border = 'rgba(148, 163, 184, 0.32)'

    return f"""
<style>
.ficha-sendero-grid {{
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(260px, 1fr));
    gap: 1rem 1.75rem;
    margin-top: 0.5rem;
    padding: 1rem 1rem 0.9rem;
    border: 1px solid {border_color};
    border-radius: 12px;
    background: {card_bg};
}}
.ficha-campo {{ display: flex; flex-direction: column; gap: 0.25rem; min-width: 0; }}
.ficha-label {{
    font-size: 0.74rem; color: {label_color}; font-weight: 700;
    text-transform: uppercase; letter-spacing: 0.03em;
}}
.ficha-valor {{
    font-size: 1.08rem; color: {value_color}; line-height: 1.45; font-weight: 600;
    word-wrap: break-word; overflow-wrap: anywhere;
}}
</style>

"""
def render_ficha_sendero(ficha: dict, proyecto: dict):
    """Muestra la ficha sin truncar valores largos (p. ej. presupuesto)."""
    titulo_color = '#ffffff'
    titulo_bg = '#1d4ed8'
    titulo_border = '#1e40af'

    label_fecha = 'Fecha entrega' if ficha['anio'] == 2024 else 'Fecha inauguración'
    campos = [
        ('Ejecutor', ficha['ejecutor']),
        ('Presupuesto', ficha['presupuesto']),
        ('Extensión (matriz)', ficha['extension']),
        ('Beneficiarios', ficha['beneficiarios']),
        (label_fecha, ficha['fecha']),
    ]
    if proyecto.get('ubicacion'):
        campos.append(('Ubicación', proyecto['ubicacion']))
    if proyecto.get('extension'):
        campos.append(('Extensión (Excel)', proyecto['extension']))
    if proyecto.get('fecha') and str(proyecto['fecha']).strip() not in ('', 'No', 'nan'):
        campos.append(('Fecha (Excel)', proyecto['fecha']))

    celdas = ''.join(
        f'<div class="ficha-campo"><span class="ficha-label">{html.escape(lbl)}</span>'
        f'<span class="ficha-valor">{html.escape(str(valor))}</span></div>'
        for lbl, valor in campos
    )
    titulo = html.escape(f'Matriz Senderos Seguros {ficha["anio"]} — {ficha["nombre"]}')
    titulo_html = (
        f'<div style="display:inline-block; color:{titulo_color} !important; '
        f'background:{titulo_bg} !important; border:1px solid {titulo_border} !important; '
        'border-radius:10px; padding:0.45rem 0.75rem 0.5rem; '
        'font-size:1.02rem; font-weight:800; line-height:1.2; '
        'text-shadow:none; box-shadow:0 1px 0 rgba(0, 0, 0, 0.12);">'
        f'{titulo}</div>'
    )
    st.markdown(
        _ficha_css()
        + f'<div style="margin-bottom:0.75rem;">{titulo_html}</div>'
        + f'<div class="ficha-sendero-grid">{celdas}</div>',
        unsafe_allow_html=True,
    )


def generar_grafico_economia(proyectos, sector_seleccionado, records):
    alias_map = build_sector_alias_map(proyectos)
    proyecto_ref = proyectos.get(sector_seleccionado, {})
    fecha_referencia = parse_flexible_date(proyecto_ref.get('fecha'))

    clasificados = []
    control_mov = {'renovacion': 0, 'emision': 0}
    excluidos_fecha = 0
    for record in records:
        sector_resuelto = resolve_sector_name(record.get('sector_raw'), alias_map)
        if sector_resuelto != sector_seleccionado:
            continue
        if not is_after_or_equal_reference(record.get('fecha_impresion'), fecha_referencia):
            excluidos_fecha += 1
            continue

        mov = normalize_text(record.get('movimiento_raw'))
        if 'renov' in mov:
            control_mov['renovacion'] += 1
        elif 'emisi' in mov:
            control_mov['emision'] += 1

        categoria, metodo = classify_business(record, fecha_referencia)
        if categoria == 'indeterminado':
            continue
        clasificados.append({
            'categoria': categoria,
            'metodo': metodo,
            'source_label': record.get('source_label', ''),
        })

    if not clasificados:
        fig, ax = plt.subplots(figsize=(10, 4.5), dpi=120)
        ax.axis('off')
        ax.text(
            0.5, 0.5,
            'No se encontraron registros para este sector con los archivos cargados.',
            ha='center', va='center', fontsize=13,
        )
        return fig, pd.DataFrame(), fecha_referencia, excluidos_fecha, control_mov

    df = pd.DataFrame(clasificados)
    resumen = (
        df.groupby('categoria')
        .size()
        .reindex(['abierto', 'renovado'], fill_value=0)
        .reset_index(name='cantidad')
    )

    fecha_txt = proyecto_ref.get('fecha', 'Sin fecha')
    titulo_sector = sector_seleccionado
    if proyecto_ref.get('ubicacion'):
        titulo_sector = f"{sector_seleccionado} | {proyecto_ref['ubicacion']}"

    fig, ax = plt.subplots(figsize=(10.5, 6), dpi=130)
    fig.patch.set_facecolor('white')
    colores = ['#1e8449', '#1f4e79']
    etiquetas = ['Abiertos', 'Renovados']
    valores = [int(resumen.loc[resumen['categoria'] == 'abierto', 'cantidad'].iloc[0]), int(resumen.loc[resumen['categoria'] == 'renovado', 'cantidad'].iloc[0])]

    barras = ax.bar(etiquetas, valores, color=colores, width=0.55)
    ax.set_title(
        f'Comparativa económica por sector\n{titulo_sector}',
        fontsize=16, fontweight='bold', pad=18,
    )
    ax.set_ylabel('Número de negocios', fontsize=12)
    ax.set_ylim(0, max(valores + [1]) * 1.35)
    ax.grid(axis='y', linestyle='--', alpha=0.25)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    for barra, cantidad in zip(barras, valores):
        ax.text(
            barra.get_x() + barra.get_width() / 2,
            barra.get_height() + max(valores + [1]) * 0.05,
            f'{cantidad}',
            ha='center', va='bottom', fontsize=12, fontweight='bold', color='#1f2937',
        )

    referencia = 'sin fecha de referencia'
    if fecha_referencia:
        if fecha_referencia.get('precision') == 'year':
            referencia = f"año {fecha_referencia['year']}"
        elif fecha_referencia.get('precision') == 'month':
            mes = [k for k, v in MONTHS_ES.items() if v == fecha_referencia['month']][0].capitalize()
            referencia = f"{mes} {fecha_referencia['year']}"
        else:
            referencia = fecha_referencia['date'].strftime('%d/%m/%Y')

    ax.text(
        0.5, -0.18,
        f'Referencia: {fecha_txt} | Comparación aplicada con {referencia}.',
        transform=ax.transAxes, ha='center', va='top', fontsize=10, color='#64748b',
    )
    fig.tight_layout()
    return fig, resumen, fecha_referencia, excluidos_fecha, control_mov


def generar_grafico_valor_suelo(registro_proyecto, source_label):
    proyecto = registro_proyecto.get('proyecto', 'Proyecto sin nombre')
    valores = [
        registro_proyecto.get('valor_2022_2023'),
        registro_proyecto.get('valor_2024'),
        registro_proyecto.get('valor_2026'),
    ]
    etiquetas = ['2022-2023', '2024-2025', '2026-2027']
    colores = ['#64748b', '#0f766e', '#b45309']

    fig, ax = plt.subplots(figsize=(10.5, 6), dpi=130)
    fig.patch.set_facecolor('white')

    valores_plot = [0 if v is None or pd.isna(v) else float(v) for v in valores]
    barras = ax.bar(etiquetas, valores_plot, color=colores, width=0.6)
    ax.set_title(
        f'Comparativa de valor de suelo por año\n{proyecto}',
        fontsize=16, fontweight='bold', pad=18,
    )
    ax.set_ylabel('Valor de suelo promedio', fontsize=12)
    ax.grid(axis='y', linestyle='--', alpha=0.25)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    max_valor = max(valores_plot or [1])
    ax.set_ylim(0, max_valor * 1.35 if max_valor > 0 else 1)

    for barra, valor in zip(barras, valores):
        if valor is None or pd.isna(valor):
            etiqueta = 'Sin dato'
            altura = 0
        else:
            altura = float(valor)
            etiqueta = f'{altura:,.2f}'
        ax.text(
            barra.get_x() + barra.get_width() / 2,
            max(altura, 0) + max_valor * 0.05 if max_valor > 0 else 0.05,
            etiqueta,
            ha='center', va='bottom', fontsize=11, fontweight='bold', color='#1f2937',
        )

    ax.text(
        0.5, -0.16,
        f'Archivo fuente: {source_label} | Se promedian los valores de todos los sectores bajo cada PROYECTO.',
        transform=ax.transAxes, ha='center', va='top', fontsize=10, color='#64748b',
    )
    fig.tight_layout()
    return fig


def get_economia_project_filters(proyectos):
    categoria_map = {
        'Todas': None,
        'Senderos Seguros': 'Senderos Seguros',
        'Zonas Metro': 'Zonas Metro',
        'Rehabilitación de Espacio Público': 'Recuperación de espacios público',
    }
    fecha_map = {
        'Todas': 'all',
        'Con fecha de inauguración/entrega': 'with_date',
        'Sin fecha de inauguración/entrega': 'without_date',
    }
    return categoria_map, fecha_map


def filter_economia_projects(proyectos, categoria_ui, fecha_filter):
    categoria_map, _ = get_economia_project_filters(proyectos)
    categoria_real = categoria_map.get(categoria_ui)

    filtrados = {}
    for nombre, proyecto in proyectos.items():
        if categoria_real and proyecto.get('categoria') != categoria_real:
            continue
        tiene_fecha = parse_flexible_date(proyecto.get('fecha')) is not None
        if fecha_filter == 'with_date' and not tiene_fecha:
            continue
        if fecha_filter == 'without_date' and tiene_fecha:
            continue
        filtrados[nombre] = proyecto
    return filtrados


def main():
    st.set_page_config(page_title=PORTAL_TITLE, layout='wide')
    render_portal_header()
    st.markdown(
        'Explora resultados de seguridad y economía por sector, con gráficos y resúmenes interactivos.'
    )

    default_path = find_default_excel()
    with st.sidebar:
        st.subheader('Fuente de datos')
        if default_path:
            st.success(f'Archivo local: `{default_path.name}`')
            st.caption(str(default_path))
        else:
            st.warning('Sin Excel en `data/`. Coloca allí el archivo del proyecto para poder cargarlo.')
        if st.button('Recargar datos', help='Vuelve a leer el Excel tras actualizarlo en disco.'):
            load_workbook_data.clear()
            st.rerun()

    file_bytes = None
    source_label = None
    cache_path = None

    if default_path is not None:
        file_bytes = default_path.read_bytes()
        source_label = default_path.name
        cache_path = default_path

    if file_bytes is None:
        st.info(
            'Coloca el archivo `.xlsx` del proyecto en la carpeta `data/` del proyecto.'
        )
        return

    cache_key = workbook_cache_key(source_label, file_bytes, cache_path)
    try:
        categorias, proyectos = load_workbook_data(cache_key, file_bytes)
    except Exception as exc:
        st.error(f'No se pudo leer el Excel ({source_label}): {exc}')
        return

    n_senderos = sum(1 for p in proyectos.values() if p.get('categoria') == 'Senderos Seguros')
    # Se omite el mensaje de "Datos cargados" en el header por solicitud del usuario.
    tab_seguridad, tab_economia, tab_valor_suelo = st.tabs(['SEGURIDAD', 'ECONOMIA', 'VALOR DE SUELO'])

    with tab_seguridad:
        st.subheader('Resumen de seguridad por proyecto')
        st.caption(
            'Esta sección permite explorar los proyectos estratégicos de seguridad y visualizar sus indicadores, fichas e imágenes asociadas.'
        )
        st.info(
            'La vista consolida cuadros comparativos y gráficos por proyecto o por categoría para analizar incidentes y delitos en cada intervención.'
        )

        modo = st.radio('Modo', ['Por proyecto', 'Por categoría'])
        if modo == 'Por proyecto':
            proy = st.selectbox('Proyecto', sorted(list(proyectos.keys())))
        else:
            cat = st.selectbox('Categoría', list(categorias.keys()))
            proy = st.selectbox('Proyecto', sorted(categorias.get(cat, [])))

        p = proyectos[proy]
        fichas = fichas_sendero(proy)
        if fichas or es_sendero_seguro(proy, p.get('categoria', '')):
            st.subheader('Ficha del sendero seguro')
            if fichas:
                for ficha in fichas:
                    with st.container(border=True):
                        render_ficha_sendero(ficha, p)
            else:
                st.caption('Proyecto de Senderos Seguros sin ficha en las matrices 2024–2026.')

            imgs = imagenes_sendero(proy)
            if imgs['extension'] or imgs['antes_despues']:
                st.markdown('#### Imágenes del sendero')
                if imgs['extension']:
                    st.markdown('**Mapa de extensión**')
                    st.image(str(imgs['extension']), use_container_width=True)
                if imgs['antes_despues']:
                    st.markdown('**Antes y después**')
                    st.image(str(imgs['antes_despues']), use_container_width=True)
                elif imgs['extension']:
                    st.caption('En el PDF no hay diapositiva de antes/después para este sendero (solo ficha).')

        ant = st.multiselect('Años anterior', PERIODOS[:-1], default=['2024'])
        act_options = _periodos_actuales_disponibles(ant)
        act_default = ['2025'] if '2025' in act_options else act_options[:1]
        act = st.multiselect('Años actual', act_options, default=act_default)

        ant_max = max((_anio_base(anio) for anio in ant), default=None)
        if ant_max is not None and any(_anio_base(anio) <= ant_max for anio in act):
            st.warning('Los Años actual deben ser posteriores al año más alto de Años anterior.')
            act = [anio for anio in act if _anio_base(anio) > ant_max]

        col1, col2, col3 = st.columns(3)
        with col1:
            btn_tabla = st.button('📊 Generar tabla comparativa')
        with col2:
            btn_grafico = st.button('📈 Generar resumen gráfico')
        with col3:
            btn_csv = st.button('💾 Exportar CSV')

        if btn_tabla:
            if not p.get('tiene_estadisticas', True):
                st.warning('Este proyecto aún no tiene datos de incidentes/delitos en el Excel.')
            elif not ant or not act:
                st.warning('Selecciona al menos un año en ambos periodos.')
            elif set(ant) & set(act):
                st.warning('No se deben repetir años en ambos periodos.')
            elif ant and act and max(_anio_base(a) for a in act) <= max(_anio_base(a) for a in ant):
                st.warning('Los Años actual deben ser posteriores al periodo anterior.')
            else:
                st.info('Generando tabla comparativa...')
                fig = generar_tabla(proyectos[proy], ant, act)
                mostrar_figura_alta_res(fig, dpi_pantalla=200)
                buf = io.BytesIO()
                fig.savefig(buf, format='png', dpi=300, bbox_inches='tight', facecolor='white')
                buf.seek(0)
                st.download_button('⬇️ Descargar tabla (PNG)', data=buf.getvalue(),
                                  file_name=f'{proy}_tabla_comparativa.png', mime='image/png')
                plt.close(fig)

        if btn_grafico:
            if not p.get('tiene_estadisticas', True):
                st.warning('Este proyecto aún no tiene datos de incidentes/delitos en el Excel.')
            elif not ant or not act:
                st.warning('Selecciona al menos un año en ambos periodos.')
            elif set(ant) & set(act):
                st.warning('No se deben repetir años en ambos periodos.')
            elif ant and act and max(_anio_base(a) for a in act) <= max(_anio_base(a) for a in ant):
                st.warning('Los Años actual deben ser posteriores al periodo anterior.')
            else:
                st.info('Generando resumen gráfico...')
                fig = generar_grafico_resumen(proyectos[proy], ant, act)
                mostrar_figura_alta_res(fig, dpi_pantalla=220)
                buf = io.BytesIO()
                fig.savefig(buf, format='png', dpi=300, bbox_inches='tight', facecolor='white')
                buf.seek(0)
                st.download_button('⬇️ Descargar gráfico (PNG)', data=buf.getvalue(),
                                  file_name=f'{proy}_grafico_resumen.png', mime='image/png')
                plt.close(fig)

        if btn_csv:
            df = export_csv(proyectos)
            csv = df.to_csv(index=False, encoding='utf-8-sig')
            st.download_button('⬇️ Descargar CSV', data=csv, file_name='seguridad_proyectos_powerbi.csv', mime='text/csv')

    with tab_economia:
        st.subheader('Conteo económico por sector')
        st.caption(
            'Esta sección resume el comportamiento económico por sector y muestra cuántos negocios se abrieron y cuántos renovaron en torno a la inauguración o entrega de cada proyecto.'
        )
        st.info(
            'La vista cruza la fecha de impresión de los registros económicos con la fecha de entrega o inauguración del proyecto para identificar aperturas y renovaciones; cuando la referencia solo tiene año, se toma ese año como base.'
        )

        economia_sources = []
        default_economia_paths = find_default_economia_excels()
        if default_economia_paths:
            for path in default_economia_paths:
                file_bytes = path.read_bytes()
                cache_key = workbook_cache_key(path.name, file_bytes, path)
                economia_sources.append({
                    'label': path.name,
                    'data': load_economia_data(cache_key, file_bytes, path.name),
                })

        if not economia_sources:
            st.warning(
                'No encontré los archivos de economía. Coloca los Excel en la carpeta data/ del proyecto con estos nombres: resultado_cruce_predios_renovacion_v3.xlsx y resultado_cruce_predios_emision_v3.xlsx.'
            )
        else:
            etiquetas_fuente = ', '.join(src['label'] for src in economia_sources)
            # Se omite el mensaje de "Archivos cargados" en el header por solicitud del usuario.

            categoria_map, fecha_map = get_economia_project_filters(proyectos)
            categoria_ui = st.selectbox(
                'Categoría',
                list(categoria_map.keys()),
                key='categoria_economia',
                help='Filtra los proyectos por su categoría original en el Excel de Seguridad.',
            )
            fecha_ui = st.selectbox(
                'Fecha de inauguración/entrega',
                list(fecha_map.keys()),
                key='fecha_economia',
                help='Filtra proyectos con fecha cargada o sin fecha en el Excel de Seguridad.',
            )

            proyectos_filtrados = filter_economia_projects(proyectos, categoria_ui, fecha_map[fecha_ui])
            if not proyectos_filtrados:
                st.warning('No hay proyectos que coincidan con esos filtros.')
                return

            sector_options = sorted(proyectos_filtrados.keys())
            sector_seleccionado = st.selectbox('Sector', sector_options, key='sector_economia')

            records = []
            for source in economia_sources:
                records.extend(source['data']['records'])

            fig, resumen, fecha_referencia, excluidos_fecha, control_mov = generar_grafico_economia(proyectos_filtrados, sector_seleccionado, records)

            if resumen.empty:
                st.warning('No se encontraron registros compatibles para ese sector con los archivos cargados.')
            else:
                mostrar_figura_alta_res(fig, dpi_pantalla=220)
                col_a, col_b = st.columns(2)
                with col_a:
                    st.metric('Negocios abiertos', int(resumen.loc[resumen['categoria'] == 'abierto', 'cantidad'].iloc[0]))
                with col_b:
                    st.metric('Negocios renovados', int(resumen.loc[resumen['categoria'] == 'renovado', 'cantidad'].iloc[0]))

                abiertos_calc = int(resumen.loc[resumen['categoria'] == 'abierto', 'cantidad'].iloc[0])
                renovados_calc = int(resumen.loc[resumen['categoria'] == 'renovado', 'cantidad'].iloc[0])
                abiertos_mov = int(control_mov.get('emision', 0))
                renovados_mov = int(control_mov.get('renovacion', 0))

                if abiertos_calc != abiertos_mov or renovados_calc != renovados_mov:
                    st.error(
                        'Control de consistencia: diferencia entre clasificación y tipo de movimiento. '
                        f'Abiertos graficados={abiertos_calc}, por Emisión={abiertos_mov}; '
                        f'Renovados graficados={renovados_calc}, por Renovación={renovados_mov}.'
                    )
                else:
                    st.caption(
                        f'Control OK: Emisión={abiertos_mov} y Renovación={renovados_mov} coinciden con la gráfica.'
                    )

                st.dataframe(
                    resumen.rename(columns={'categoria': 'Estado', 'cantidad': 'Cantidad'}).assign(
                        Estado=lambda df: df['Estado'].map({'abierto': 'Abiertos', 'renovado': 'Renovados'})
                    ),
                    use_container_width=True,
                    hide_index=True,
                )

                if excluidos_fecha:
                    st.caption(f'Registros excluidos por estar antes de la fecha de referencia: {excluidos_fecha}')

                if fecha_referencia:
                    st.caption(
                        f"Fecha de referencia del sector: {proyectos[sector_seleccionado].get('fecha', 'Sin fecha')}"
                    )

                buf = io.BytesIO()
                fig.savefig(buf, format='png', dpi=300, bbox_inches='tight', facecolor='white')
                buf.seek(0)
                st.download_button(
                    '⬇️ Descargar gráfico económico (PNG)',
                    data=buf.getvalue(),
                    file_name=f'{sector_seleccionado}_economia.png',
                    mime='image/png'
                )
                plt.close(fig)

    with tab_valor_suelo:
        st.subheader('Valor de Suelo')
        st.caption(
            'Esta sección agrupa todos los sectores de cada PROYECTO y compara el promedio del valor de suelo para 2022-2023, 2024-2025 y 2026-2027.'
        )
        valor_suelo_bytes = None
        valor_suelo_source = None
        valor_suelo_cache_key = None

        default_valor_suelo_path = find_default_valor_suelo_excel()
        if default_valor_suelo_path is not None:
            valor_suelo_bytes = default_valor_suelo_path.read_bytes()
            valor_suelo_source = default_valor_suelo_path.name
            valor_suelo_cache_key = workbook_cache_key(valor_suelo_source, valor_suelo_bytes, default_valor_suelo_path)

        if valor_suelo_bytes is None:
            st.warning(
                'No encontré un Excel de valor de suelo. Colócalo en data/ con un nombre como AIVAS_cruce_codigos_nuevo.xlsx o define VALOR_SUELO_EXCEL.'
            )
        else:
            try:
                valor_suelo_data = load_valor_suelo_data(valor_suelo_cache_key, valor_suelo_bytes, valor_suelo_source)
            except Exception as exc:
                st.error(f'No se pudo leer el Excel de valor de suelo ({valor_suelo_source}): {exc}')
            else:
                resumen_suelo = valor_suelo_data['resumen']
                if not resumen_suelo.empty:
                    resumen_suelo = resumen_suelo.copy()
                    resumen_suelo['categoria'] = resumen_suelo['proyecto'].apply(resolve_valor_suelo_categoria)

                if resumen_suelo.empty:
                    st.warning('No se encontraron filas válidas con PROYECTO y valores de suelo en el Excel cargado.')
                else:
                    categoria_valor_suelo = st.selectbox(
                        'Categoría',
                        VALOR_SUELO_CATEGORIA_ORDER,
                        key='categoria_valor_suelo',
                    )

                    resumen_filtrado = filtrar_valor_suelo_por_categoria(resumen_suelo, categoria_valor_suelo)
                    if resumen_filtrado.empty:
                        st.warning('No hay proyectos para esa categoría.')
                    else:
                        st.caption(
                            f'Categoría seleccionada: **{categoria_valor_suelo}** · {len(resumen_filtrado)} proyectos disponibles'
                        )

                        proyecto_valor_suelo = st.selectbox(
                            'Proyecto',
                            resumen_filtrado['proyecto'].tolist(),
                            key='proyecto_valor_suelo',
                        )

                        fila_proyecto = resumen_filtrado.loc[resumen_filtrado['proyecto'] == proyecto_valor_suelo].iloc[0]
                        fig = generar_grafico_valor_suelo(fila_proyecto, valor_suelo_source)
                        mostrar_figura_alta_res(fig, dpi_pantalla=220)

                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric('2022-2023', 'Sin dato' if pd.isna(fila_proyecto['valor_2022_2023']) else f"{fila_proyecto['valor_2022_2023']:,.2f}")
                        with col2:
                            st.metric('2024-2025', 'Sin dato' if pd.isna(fila_proyecto['valor_2024']) else f"{fila_proyecto['valor_2024']:,.2f}")
                        with col3:
                            st.metric('2026-2027', 'Sin dato' if pd.isna(fila_proyecto['valor_2026']) else f"{fila_proyecto['valor_2026']:,.2f}")

                        detalle_proyecto = pd.DataFrame(valor_suelo_data['records'])
                        detalle_proyecto = detalle_proyecto[detalle_proyecto['proyecto_raw'] == fila_proyecto['proyecto']]
                        if not detalle_proyecto.empty:
                            st.markdown('#### Sectores utilizados en el promedio')
                            st.dataframe(
                                detalle_proyecto[['descripcion_raw', 'valor_2022_2023', 'valor_2024', 'valor_2026']].rename(
                                    columns={
                                        'descripcion_raw': 'DESCRIPCION',
                                        'valor_2022_2023': '2022-2023',
                                        'valor_2024': '2024-2025',
                                        'valor_2026': '2026-2027',
                                    }
                                ),
                                use_container_width=True,
                                hide_index=True,
                            )

                        buf = io.BytesIO()
                        fig.savefig(buf, format='png', dpi=300, bbox_inches='tight', facecolor='white')
                        buf.seek(0)
                        st.download_button(
                            '⬇️ Descargar gráfico de valor de suelo (PNG)',
                            data=buf.getvalue(),
                            file_name=f'{proyecto_valor_suelo}_valor_de_suelo.png',
                            mime='image/png',
                        )
                        plt.close(fig)


if __name__ == '__main__':
    main()
